#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
pytest共通フィクスチャ
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, AsyncMock

import pytest

# プロジェクトルートをパスに追加
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


@pytest.fixture
def sample_player_data():
    """テスト用プレイヤーデータ"""
    from core.excel_handler import PlayerData

    return [
        PlayerData(
            row_index=2,
            player_name="楽天カード",
            official_url="https://www.rakuten-card.co.jp/",
            company_name="楽天カード株式会社",
        ),
        PlayerData(
            row_index=3,
            player_name="三井住友カード",
            official_url="https://www.smbc-card.com/",
            company_name="三井住友カード株式会社",
        ),
        PlayerData(
            row_index=4,
            player_name="JCBカード",
            official_url="https://www.jcb.co.jp/",
            company_name="株式会社ジェーシービー",
        ),
    ]


@pytest.fixture
def mock_llm_client():
    """モック化されたLLMクライアント"""
    mock = MagicMock()
    mock.call.return_value = '''```json
{
    "is_active": true,
    "change_type": "none",
    "current_service_name": "楽天カード",
    "current_company_name": "楽天カード株式会社",
    "current_url": "https://www.rakuten-card.co.jp/",
    "changes": [],
    "news": "",
    "confidence": 0.95,
    "sources": ["https://www.rakuten-card.co.jp/"]
}
```'''
    mock.extract_json.return_value = {
        "is_active": True,
        "change_type": "none",
        "current_service_name": "楽天カード",
        "current_company_name": "楽天カード株式会社",
        "current_url": "https://www.rakuten-card.co.jp/",
        "changes": [],
        "news": "",
        "confidence": 0.95,
        "sources": ["https://www.rakuten-card.co.jp/"]
    }
    return mock


@pytest.fixture
def mock_llm_client_withdrawal():
    """撤退ケース用のモックLLMクライアント"""
    mock = MagicMock()
    mock.call.return_value = '''```json
{
    "is_active": false,
    "change_type": "withdrawal",
    "current_service_name": "終了したサービス",
    "current_company_name": "株式会社〇〇",
    "current_url": "",
    "changes": ["2025年3月にサービス終了"],
    "news": "2025年3月31日をもってサービスを終了",
    "confidence": 0.9,
    "sources": ["https://example.com/press-release"]
}
```'''
    mock.extract_json.return_value = {
        "is_active": False,
        "change_type": "withdrawal",
        "current_service_name": "終了したサービス",
        "current_company_name": "株式会社〇〇",
        "current_url": "",
        "changes": ["2025年3月にサービス終了"],
        "news": "2025年3月31日をもってサービスを終了",
        "confidence": 0.9,
        "sources": ["https://example.com/press-release"]
    }
    return mock


@pytest.fixture
def temp_excel_file(tmp_path):
    """テスト用一時Excelファイルを作成"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # ヘッダー行（4行目を想定）
    headers = ["調査票用No", "サービス名", "事業者名", "公式URL", "備考"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)

    # データ行
    data = [
        [1, "テストサービスA", "テスト株式会社A", "https://test-a.example.com/", ""],
        [2, "テストサービスB", "テスト株式会社B", "https://test-b.example.com/", "備考テスト"],
        [3, "テストサービスC", "テスト株式会社C", "", "URLなし"],
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    file_path = tmp_path / "test_player_list.xlsx"
    wb.save(file_path)

    return file_path


# ====================================
# 店舗調査用フィクスチャ
# ====================================
@pytest.fixture
def sample_store_companies():
    """テスト用企業リスト"""
    return [
        {
            "company_name": "テスト株式会社A",
            "official_url": "https://test-a.example.com/",
            "industry": "飲食店",
        },
        {
            "company_name": "テスト株式会社B",
            "official_url": "https://test-b.example.com/",
            "industry": "飲食店",
        },
        {
            "company_name": "テスト株式会社C",
            "official_url": "",
            "industry": "小売店",
        },
    ]


@pytest.fixture
def mock_llm_client_store_success():
    """店舗調査成功ケース用のモックLLMクライアント"""
    mock = MagicMock()
    mock.call.return_value = '''```json
{
    "total_stores": 100,
    "direct_stores": 80,
    "franchise_stores": 20,
    "prefecture_distribution": {
        "東京都": 20,
        "大阪府": 15,
        "愛知県": 10
    },
    "confidence": 0.85,
    "sources": ["https://example.com/stores/", "https://example.com/ir/report.pdf"],
    "notes": "2026年1月時点のIR資料より"
}
```'''
    return mock


@pytest.fixture
def sample_store_investigation_results():
    """テスト用店舗調査結果"""
    from datetime import datetime
    from investigators.base import StoreInvestigationResult

    return [
        StoreInvestigationResult(
            company_name="テスト株式会社A",
            total_stores=100,
            source_urls=["https://test-a.example.com/stores/"],
            investigation_date=datetime.now(),
            investigation_mode="ai",
            direct_stores=80,
            franchise_stores=20,
            prefecture_distribution={"東京都": 20, "大阪府": 15},
            notes="調査完了",
            needs_verification=False,
        ),
        StoreInvestigationResult(
            company_name="テスト株式会社B",
            total_stores=50,
            source_urls=["https://test-b.example.com/"],
            investigation_date=datetime.now(),
            investigation_mode="ai",
            notes="要確認",
            needs_verification=True,
        ),
    ]
