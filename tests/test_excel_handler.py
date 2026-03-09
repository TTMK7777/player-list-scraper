#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel Handler のテスト
"""

import pytest
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.excel_handler import ExcelHandler, PlayerData, ValidationReportExporter


class TestExcelHandler:
    """ExcelHandler のテストクラス"""

    def test_load_excel_success(self, temp_excel_file):
        """正常なExcelファイルの読み込み"""
        handler = ExcelHandler()
        players = handler.load(temp_excel_file)

        assert len(players) == 3
        assert players[0].player_name == "テストサービスA"
        assert players[0].company_name == "テスト株式会社A"
        assert players[0].official_url == "https://test-a.example.com/"

    def test_load_excel_detects_header_row(self, temp_excel_file):
        """ヘッダー行の自動検出"""
        handler = ExcelHandler()
        handler.load(temp_excel_file)

        # ヘッダー行が4行目であることを確認
        assert handler.header_row == 4

    def test_load_excel_detects_columns(self, temp_excel_file):
        """列の自動検出"""
        handler = ExcelHandler()
        handler.load(temp_excel_file)

        column_names = handler.get_column_names()
        assert "サービス名" in column_names
        assert "事業者名" in column_names
        assert "公式URL" in column_names

    def test_load_nonexistent_file(self):
        """存在しないファイルの読み込み"""
        handler = ExcelHandler()

        with pytest.raises(FileNotFoundError):
            handler.load("/nonexistent/path/file.xlsx")

    def test_player_data_with_empty_url(self, temp_excel_file):
        """URLが空のプレイヤーデータ"""
        handler = ExcelHandler()
        players = handler.load(temp_excel_file)

        # 3番目のデータはURLが空
        player_c = [p for p in players if p.player_name == "テストサービスC"][0]
        assert player_c.official_url == ""

    def test_player_data_row_index(self, temp_excel_file):
        """行インデックスが正しく保持されているか"""
        handler = ExcelHandler()
        players = handler.load(temp_excel_file)

        # データ行は5行目から開始
        assert players[0].row_index == 5
        assert players[1].row_index == 6
        assert players[2].row_index == 7


class TestExcelHandlerFallback:
    """Excel列検出フォールバックのテスト"""

    def test_no_column_skips_numeric_first_column(self, tmp_path):
        """No.列(列1)+プレイヤー名列(列2)のExcelで列2が自動選択される"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        # ヘッダー行（プレイヤー名パターンにマッチしない列名）
        ws.cell(row=1, column=1, value="No.")
        ws.cell(row=1, column=2, value="対象")

        # データ行
        ws.cell(row=2, column=1, value="1")
        ws.cell(row=2, column=2, value="Netflix")
        ws.cell(row=3, column=1, value="2")
        ws.cell(row=3, column=2, value="Hulu")
        ws.cell(row=4, column=1, value="3")
        ws.cell(row=4, column=2, value="SAMANSA")

        file_path = tmp_path / "no_col_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        # 列2がフォールバックとして選択され、全プレイヤーが読み込まれる
        assert len(players) == 3
        names = [p.player_name for p in players]
        assert "Netflix" in names
        assert "Hulu" in names
        assert "SAMANSA" in names
        # 列1（No.列）ではなく列2が_player_nameに設定されていること
        assert handler.column_map.get("_player_name") == 2

    def test_double_numeric_columns_skipped(self, tmp_path):
        """No.列(列1)+No.列(列2)+プレイヤー名列(列3)のExcelで列3が選択される"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="No.")
        ws.cell(row=1, column=2, value="管理番号")
        ws.cell(row=1, column=3, value="対象")

        ws.cell(row=2, column=1, value="1")
        ws.cell(row=2, column=2, value="100")
        ws.cell(row=2, column=3, value="Netflix")
        ws.cell(row=3, column=1, value="2")
        ws.cell(row=3, column=2, value="200")
        ws.cell(row=3, column=3, value="Hulu")

        file_path = tmp_path / "double_no_col_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        assert len(players) == 2
        assert handler.column_map.get("_player_name") == 3
        assert players[0].player_name == "Netflix"
        assert players[1].player_name == "Hulu"

    def test_all_numeric_columns_fallback_to_col1(self, tmp_path):
        """全列が数字のみの場合、列1を最終フォールバックとして使用"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="コードA")
        ws.cell(row=1, column=2, value="コードB")

        ws.cell(row=2, column=1, value="100")
        ws.cell(row=2, column=2, value="200")
        ws.cell(row=3, column=1, value="300")
        ws.cell(row=3, column=2, value="400")

        file_path = tmp_path / "all_numeric_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        # 全列数字のため列1にフォールバック、数字のプレイヤー名もスキップされない
        assert handler.column_map.get("_player_name") == 1
        assert len(players) == 2
        assert players[0].player_name == "100"
        assert players[1].player_name == "300"

    def test_samansa_bug_scenario(self, tmp_path):
        """SAMANSAバグ再現: 定額制動画配信でNo.7 SAMANSAが反映される"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        # 実際のプレイヤーリスト構造を模擬
        ws.cell(row=1, column=1, value="No.")
        ws.cell(row=1, column=2, value="対象サービス")

        services = [
            "Netflix", "Amazonプライム・ビデオ", "U-NEXT", "Hulu",
            "Disney+", "dアニメストア", "SAMANSA", "ABEMA",
        ]
        for i, name in enumerate(services, 1):
            ws.cell(row=i + 1, column=1, value=str(i))
            ws.cell(row=i + 1, column=2, value=name)

        file_path = tmp_path / "samansa_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        names = [p.player_name for p in players]
        assert "SAMANSA" in names, "SAMANSAが読み込まれていない"
        assert len(players) == 8

    def test_header_detection_meishou(self, tmp_path):
        """新パターン「名称」でのヘッダー検出テスト"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="名称")
        ws.cell(row=1, column=2, value="備考")
        ws.cell(row=2, column=1, value="テストサービス")
        ws.cell(row=2, column=2, value="メモ")

        file_path = tmp_path / "meishou_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        assert handler.column_map.get("_player_name") == 1
        assert len(players) == 1
        assert players[0].player_name == "テストサービス"

    def test_header_detection_service(self, tmp_path):
        """新パターン「サービス」でのヘッダー検出テスト"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="サービス")
        ws.cell(row=2, column=1, value="Hulu")

        file_path = tmp_path / "service_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        assert handler.column_map.get("_player_name") == 1
        assert players[0].player_name == "Hulu"

    def test_header_detection_chousataishou(self, tmp_path):
        """新パターン「調査対象」でのヘッダー検出テスト"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="調査対象")
        ws.cell(row=2, column=1, value="楽天モバイル")

        file_path = tmp_path / "chousataishou_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        assert handler.column_map.get("_player_name") == 1
        assert players[0].player_name == "楽天モバイル"

    def test_header_detection_name_english(self, tmp_path):
        """新パターン「name」でのヘッダー検出テスト"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="name")
        ws.cell(row=2, column=1, value="TestPlayer")

        file_path = tmp_path / "name_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        players = handler.load(file_path)

        assert handler.column_map.get("_player_name") == 1
        assert players[0].player_name == "TestPlayer"

    def test_warnings_attribute_accumulation(self, tmp_path):
        """warnings属性にフォールバック警告が蓄積される"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="No.")
        ws.cell(row=1, column=2, value="対象")
        ws.cell(row=2, column=1, value="1")
        ws.cell(row=2, column=2, value="テストA")

        file_path = tmp_path / "warnings_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        assert handler.warnings == []  # 初期状態は空

        handler.load(file_path)

        # フォールバック使用時に警告が蓄積される
        assert len(handler.warnings) == 1
        assert "フォールバック" in handler.warnings[0]

    def test_fallback_warning_message_content(self, tmp_path):
        """フォールバック警告メッセージに列番号と列名が含まれる"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="No.")
        ws.cell(row=1, column=2, value="対象名")
        ws.cell(row=2, column=1, value="1")
        ws.cell(row=2, column=2, value="テストサービス")

        file_path = tmp_path / "warning_content_test.xlsx"
        wb.save(file_path)

        handler = ExcelHandler()
        handler.load(file_path)

        assert len(handler.warnings) == 1
        warning_msg = handler.warnings[0]
        assert "列2" in warning_msg
        assert "対象名" in warning_msg
        assert "自動検出されませんでした" in warning_msg

    def test_fallback_warning_logged(self, tmp_path, caplog):
        """フォールバック列使用時にwarningがログ出力される"""
        import openpyxl
        import logging

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="列A")
        ws.cell(row=2, column=1, value="テストプレイヤー")

        file_path = tmp_path / "fallback_warning_test.xlsx"
        wb.save(file_path)

        with caplog.at_level(logging.WARNING):
            handler = ExcelHandler()
            handler.load(file_path)

        assert any("フォールバック" in record.message for record in caplog.records)


class TestValidationReportExporter:
    """ValidationReportExporter のテストクラス"""

    def test_export_creates_file(self, tmp_path):
        """エクスポートがファイルを作成するか"""
        from investigators.base import (
            AlertLevel, ChangeType, ValidationStatus, ValidationResult
        )
        from datetime import datetime

        results = [
            ValidationResult(
                player_name_original="テストサービス",
                player_name_current="テストサービス",
                status=ValidationStatus.UNCHANGED,
                alert_level=AlertLevel.OK,
                change_type=ChangeType.NO_CHANGE,
                url_original="https://example.com",
                url_current="https://example.com",
                checked_at=datetime.now(),
            )
        ]

        exporter = ValidationReportExporter()
        output_path = tmp_path / "test_report.xlsx"
        result_path = exporter.export(results, output_path)

        assert result_path.exists()

    def test_export_with_critical_alert(self, tmp_path):
        """緊急アラートのエクスポート"""
        from investigators.base import (
            AlertLevel, ChangeType, ValidationStatus, ValidationResult
        )
        from datetime import datetime

        results = [
            ValidationResult(
                player_name_original="撤退サービス",
                player_name_current="撤退サービス",
                status=ValidationStatus.CONFIRMED,
                alert_level=AlertLevel.CRITICAL,
                change_type=ChangeType.WITHDRAWAL,
                change_details=["2025年3月にサービス終了"],
                url_original="https://example.com",
                url_current="",
                needs_manual_review=False,
                checked_at=datetime.now(),
            )
        ]

        exporter = ValidationReportExporter()
        output_path = tmp_path / "test_critical_report.xlsx"
        result_path = exporter.export(results, output_path)

        assert result_path.exists()

        # ファイルを読み込んで内容を確認
        import openpyxl
        wb = openpyxl.load_workbook(result_path)
        ws = wb.active

        # 2行目（データ行）のアラートレベルを確認
        assert "緊急" in ws.cell(row=2, column=1).value
