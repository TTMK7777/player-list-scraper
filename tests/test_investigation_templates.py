#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
調査テンプレート管理 (investigation_templates) のテスト
"""

import sys
import json
from pathlib import Path
from datetime import datetime

import pytest

# プロジェクトルートをパスに追加
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core.investigation_templates import (
    InvestigationTemplate,
    TemplateManager,
    VALID_CATEGORIES,
)


# ====================================
# InvestigationTemplate テスト
# ====================================
class TestInvestigationTemplate:
    """調査テンプレートデータクラスのテスト"""

    def test_create_template(self):
        """正常なテンプレート作成"""
        template = InvestigationTemplate(
            id="test_template",
            label="テストテンプレート",
            description="テスト用のテンプレートです",
            category="属性系",
            attributes=["属性A", "属性B", "属性C"],
            context="判定基準: 属性A=有料サービス",
        )
        assert template.id == "test_template"
        assert template.label == "テストテンプレート"
        assert template.category == "属性系"
        assert len(template.attributes) == 3
        assert template.context == "判定基準: 属性A=有料サービス"
        assert template.is_builtin is False

    def test_template_validation_empty_id(self):
        """IDが空の場合にValueError"""
        with pytest.raises(ValueError, match="テンプレートIDは必須"):
            InvestigationTemplate(
                id="",
                label="テスト",
                category="属性系",
                attributes=["属性A"],
                context="",
            )

    def test_template_validation_empty_label(self):
        """ラベルが空の場合にValueError"""
        with pytest.raises(ValueError, match="テンプレートラベルは必須"):
            InvestigationTemplate(
                id="test",
                label="",
                category="属性系",
                attributes=["属性A"],
                context="",
            )

    def test_template_validation_invalid_category(self):
        """無効なカテゴリでValueError"""
        with pytest.raises(ValueError, match="無効なカテゴリ"):
            InvestigationTemplate(
                id="test",
                label="テスト",
                category="無効なカテゴリ",
                attributes=["属性A"],
                context="",
            )

    def test_template_validation_empty_attributes(self):
        """属性リストが空の場合にValueError"""
        with pytest.raises(ValueError, match="属性リストは1つ以上必要"):
            InvestigationTemplate(
                id="test",
                label="テスト",
                category="属性系",
                attributes=[],
                context="",
            )

    def test_template_to_dict_roundtrip(self):
        """to_dict → from_dict のラウンドトリップ"""
        original = InvestigationTemplate(
            id="roundtrip_test",
            label="ラウンドトリップテスト",
            description="往復確認用",
            category="分類系",
            attributes=["A", "B", "C"],
            context="コンテキスト情報",
            batch_size=7,
            is_builtin=False,
        )

        data = original.to_dict()
        restored = InvestigationTemplate.from_dict(data)

        assert restored.id == original.id
        assert restored.label == original.label
        assert restored.description == original.description
        assert restored.category == original.category
        assert restored.attributes == original.attributes
        assert restored.context == original.context
        assert restored.batch_size == original.batch_size
        assert restored.is_builtin == original.is_builtin

    def test_builtin_flag(self):
        """is_builtin フラグの動作確認"""
        template_user = InvestigationTemplate(
            id="user_template",
            label="ユーザーテンプレート",
            category="カスタム",
            attributes=["属性1"],
            context="",
            is_builtin=False,
        )
        template_builtin = InvestigationTemplate(
            id="builtin_template",
            label="組み込みテンプレート",
            category="属性系",
            attributes=["属性1"],
            context="",
            is_builtin=True,
        )
        assert template_user.is_builtin is False
        assert template_builtin.is_builtin is True

    def test_area_template_attributes(self):
        """地理系テンプレートの属性数が正しいか（8地方=8, 47都道府県=47）"""
        # 8地方
        template_8 = InvestigationTemplate(
            id="地方区分_8地方",
            label="8地方区分",
            category="地理系",
            attributes=["北海道", "東北", "関東", "中部", "近畿", "中国", "四国", "九州・沖縄"],
            context="地方区分コンテキスト",
        )
        assert len(template_8.attributes) == 8

        # 47都道府県（実際の47個を列挙）
        prefectures_47 = [
            "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
            "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
            "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
            "岐阜県", "静岡県", "愛知県", "三重県",
            "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
            "鳥取県", "島根県", "岡山県", "広島県", "山口県",
            "徳島県", "香川県", "愛媛県", "高知県",
            "福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
        ]
        template_47 = InvestigationTemplate(
            id="都道府県_47",
            label="47都道府県",
            category="地理系",
            attributes=prefectures_47,
            context="都道府県コンテキスト",
        )
        assert len(template_47.attributes) == 47

    def test_template_with_context(self):
        """context フィールドが正しく保持されるか"""
        context_text = "属性A=有料サービス, 属性B=無料サービス"
        template = InvestigationTemplate(
            id="context_test",
            label="コンテキストテスト",
            category="属性系",
            attributes=["属性A", "属性B"],
            context=context_text,
        )
        assert template.context == context_text

        # to_dict でも保持されるか
        data = template.to_dict()
        assert data["context"] == context_text

        # from_dict でも復元されるか
        restored = InvestigationTemplate.from_dict(data)
        assert restored.context == context_text

    def test_from_dict_missing_required_fields(self):
        """必須フィールド不足時にValueError"""
        with pytest.raises(ValueError, match="必須フィールド"):
            InvestigationTemplate.from_dict({
                "id": "test",
                # label がない
                "category": "属性系",
                "attributes": ["A"],
            })

    def test_from_dict_default_values(self):
        """from_dict でデフォルト値が適用されるか"""
        minimal_data = {
            "id": "minimal",
            "label": "最小構成",
            "category": "カスタム",
            "attributes": ["属性X"],
        }
        template = InvestigationTemplate.from_dict(minimal_data)
        assert template.description == ""
        assert template.context == ""
        assert template.batch_size is None
        assert template.is_builtin is False
        assert template.created_at is not None
        assert template.updated_at is not None


# ====================================
# TemplateManager テスト
# ====================================
class TestTemplateManager:
    """テンプレートマネージャーのテスト"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """テスト用の一時ディレクトリ"""
        builtin_dir = tmp_path / "templates" / "builtin"
        user_dir = tmp_path / "templates" / "user"
        builtin_dir.mkdir(parents=True, exist_ok=True)
        user_dir.mkdir(parents=True, exist_ok=True)

        # テスト用組み込みテンプレートを作成
        builtin_template = {
            "id": "builtin_test",
            "label": "組み込みテスト",
            "description": "テスト用組み込みテンプレート",
            "category": "属性系",
            "attributes": ["属性1", "属性2"],
            "context": "組み込みコンテキスト",
            "batch_size": 5,
            "is_builtin": True,
            "created_at": "2026-02-17T00:00:00",
            "updated_at": "2026-02-17T00:00:00",
        }
        with open(builtin_dir / "builtin_test.json", "w", encoding="utf-8") as f:
            json.dump(builtin_template, f, ensure_ascii=False)

        return tmp_path

    def test_list_all_templates(self, temp_dir):
        """全テンプレートのリスト取得"""
        manager = TemplateManager(project_root=temp_dir)
        templates = manager.list_templates()
        assert len(templates) >= 1
        assert any(t.id == "builtin_test" for t in templates)

    def test_list_by_category(self, temp_dir):
        """カテゴリフィルタ"""
        manager = TemplateManager(project_root=temp_dir)

        # ユーザーテンプレート追加（カテゴリ: 地理系）
        user_template = InvestigationTemplate(
            id="user_geo",
            label="ユーザー地理系",
            category="地理系",
            attributes=["地域A", "地域B"],
            context="",
        )
        manager.save_template(user_template)

        # 属性系のみ取得
        attr_templates = manager.list_templates(category="属性系")
        assert all(t.category == "属性系" for t in attr_templates)
        assert any(t.id == "builtin_test" for t in attr_templates)

        # 地理系のみ取得
        geo_templates = manager.list_templates(category="地理系")
        assert all(t.category == "地理系" for t in geo_templates)
        assert any(t.id == "user_geo" for t in geo_templates)

    def test_get_builtin_template(self, temp_dir):
        """組み込みテンプレートの取得"""
        manager = TemplateManager(project_root=temp_dir)
        template = manager.get_template("builtin_test")
        assert template.id == "builtin_test"
        assert template.label == "組み込みテスト"
        assert template.is_builtin is True

    def test_save_and_load_user_template(self, temp_dir):
        """ユーザーテンプレートの保存と読み込み"""
        manager = TemplateManager(project_root=temp_dir)

        # 保存
        user_template = InvestigationTemplate(
            id="user_save_test",
            label="ユーザー保存テスト",
            description="保存テスト用",
            category="カスタム",
            attributes=["属性X", "属性Y", "属性Z"],
            context="保存テストコンテキスト",
            batch_size=3,
            is_builtin=False,
        )
        save_path = manager.save_template(user_template)
        assert save_path.exists()
        assert save_path.suffix == ".json"

        # 読み込み
        loaded = manager.get_template("user_save_test")
        assert loaded.id == "user_save_test"
        assert loaded.label == "ユーザー保存テスト"
        assert loaded.description == "保存テスト用"
        assert loaded.category == "カスタム"
        assert len(loaded.attributes) == 3
        assert loaded.context == "保存テストコンテキスト"
        assert loaded.batch_size == 3
        assert loaded.is_builtin is False

    def test_delete_user_template(self, temp_dir):
        """ユーザーテンプレートの削除"""
        manager = TemplateManager(project_root=temp_dir)

        # ユーザーテンプレート作成
        user_template = InvestigationTemplate(
            id="user_delete_test",
            label="削除テスト",
            category="カスタム",
            attributes=["属性1"],
            context="",
        )
        manager.save_template(user_template)

        # 削除前は取得できる
        assert manager.get_template("user_delete_test") is not None

        # 削除
        result = manager.delete_template("user_delete_test")
        assert result is True

        # 削除後は取得できない
        with pytest.raises(KeyError):
            manager.get_template("user_delete_test")

    def test_delete_builtin_blocked(self, temp_dir):
        """組み込みテンプレートの削除がブロックされる"""
        manager = TemplateManager(project_root=temp_dir)

        with pytest.raises(PermissionError, match="組み込みテンプレート"):
            manager.delete_template("builtin_test")

    def test_save_builtin_blocked(self, temp_dir):
        """組み込みフラグがTrueのテンプレート保存がブロックされる"""
        manager = TemplateManager(project_root=temp_dir)

        builtin_template = InvestigationTemplate(
            id="should_fail",
            label="保存不可",
            category="属性系",
            attributes=["属性1"],
            context="",
            is_builtin=True,  # これが原因でエラー
        )

        with pytest.raises(PermissionError, match="組み込みテンプレートは保存できません"):
            manager.save_template(builtin_template)

    def test_import_from_text_comma(self):
        """カンマ区切りテキストからのインポート"""
        manager = TemplateManager()
        text = "属性A, 属性B, 属性C"
        result = manager.import_from_text(text, separator=",")
        assert len(result) == 3
        assert result == ["属性A", "属性B", "属性C"]

    def test_import_from_text_newline(self):
        """改行区切りテキストからのインポート"""
        manager = TemplateManager()
        text = "属性A\n属性B\n属性C"
        result = manager.import_from_text(text, separator=",")
        assert len(result) == 3
        assert result == ["属性A", "属性B", "属性C"]

    def test_import_from_text_mixed(self):
        """カンマと改行の混在テキストからのインポート"""
        manager = TemplateManager()
        text = "属性A, 属性B\n属性C, 属性D"
        result = manager.import_from_text(text, separator=",")
        assert len(result) == 4
        assert result == ["属性A", "属性B", "属性C", "属性D"]

    def test_import_from_text_with_empty_items(self):
        """空白項目を除外してインポート"""
        manager = TemplateManager()
        text = "属性A, , 属性B,  , 属性C"
        result = manager.import_from_text(text, separator=",")
        assert len(result) == 3
        assert result == ["属性A", "属性B", "属性C"]

    def test_import_from_excel(self, tmp_path):
        """Excelファイルからのインポート"""
        # openpyxl をインポート（未インストールならスキップ）
        pytest.importorskip("openpyxl")

        import openpyxl

        # テスト用Excelファイルを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "属性A"
        ws["A2"] = "属性B"
        ws["A3"] = "属性C"
        ws["A4"] = ""  # 空セル
        ws["A5"] = "属性D"

        excel_path = tmp_path / "test_attributes.xlsx"
        wb.save(excel_path)
        wb.close()

        # インポート
        manager = TemplateManager()
        result = manager.import_from_excel(excel_path, column=0)

        assert len(result) == 4  # 空セルは除外
        assert result == ["属性A", "属性B", "属性C", "属性D"]

    def test_import_from_excel_second_column(self, tmp_path):
        """Excelファイルから第2列をインポート"""
        pytest.importorskip("openpyxl")

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "無視"
        ws["B1"] = "属性X"
        ws["A2"] = "無視"
        ws["B2"] = "属性Y"

        excel_path = tmp_path / "test_second_column.xlsx"
        wb.save(excel_path)
        wb.close()

        manager = TemplateManager()
        result = manager.import_from_excel(excel_path, column=1)

        assert len(result) == 2
        assert result == ["属性X", "属性Y"]

    def test_import_from_excel_not_found(self):
        """存在しないExcelファイルでFileNotFoundError"""
        manager = TemplateManager()

        with pytest.raises(FileNotFoundError):
            manager.import_from_excel(Path("non_existent.xlsx"))

    def test_duplicate_id_handling(self, temp_dir):
        """同じIDで保存した場合の上書き動作"""
        manager = TemplateManager(project_root=temp_dir)

        # 初回保存
        template_v1 = InvestigationTemplate(
            id="overwrite_test",
            label="初回バージョン",
            category="カスタム",
            attributes=["属性1"],
            context="初回コンテキスト",
        )
        manager.save_template(template_v1)

        # 同じIDで再保存（上書き）
        template_v2 = InvestigationTemplate(
            id="overwrite_test",
            label="上書きバージョン",
            category="カスタム",
            attributes=["属性1", "属性2", "属性3"],
            context="上書きコンテキスト",
        )
        manager.save_template(template_v2)

        # 読み込んで確認（v2 の内容になっているはず）
        loaded = manager.get_template("overwrite_test")
        assert loaded.label == "上書きバージョン"
        assert len(loaded.attributes) == 3
        assert loaded.context == "上書きコンテキスト"

    def test_invalid_template_validation(self):
        """不正なテンプレートデータでのエラー"""
        manager = TemplateManager()

        # 不正なテンプレート（無効なカテゴリ）
        invalid_template = InvestigationTemplate(
            id="will_fail_on_init",
            label="不正テンプレート",
            category="カスタム",
            attributes=["属性1"],
            context="",
        )

        # バリデーションは __post_init__ で行われるため、
        # from_dict で無効なデータを読み込むとエラー
        with pytest.raises(ValueError):
            InvestigationTemplate.from_dict({
                "id": "invalid",
                "label": "不正",
                "category": "存在しないカテゴリ",
                "attributes": ["A"],
            })

    def test_export_template_json(self, temp_dir):
        """テンプレートのJSON エクスポート"""
        manager = TemplateManager(project_root=temp_dir)

        # エクスポート
        exported_data = manager.export_template("builtin_test")

        assert isinstance(exported_data, dict)
        assert exported_data["id"] == "builtin_test"
        assert exported_data["label"] == "組み込みテスト"
        assert exported_data["category"] == "属性系"
        assert exported_data["is_builtin"] is True

    def test_export_template_not_found(self):
        """存在しないテンプレートのエクスポートでKeyError"""
        manager = TemplateManager()

        with pytest.raises(KeyError):
            manager.export_template("non_existent_template")

    def test_get_template_not_found(self):
        """存在しないテンプレートの取得でKeyError"""
        manager = TemplateManager()

        with pytest.raises(KeyError, match="テンプレート .* が見つかりません"):
            manager.get_template("definitely_does_not_exist")

    def test_delete_template_not_found(self, temp_dir):
        """存在しないテンプレートの削除でKeyError"""
        manager = TemplateManager(project_root=temp_dir)

        with pytest.raises(KeyError, match="テンプレート .* が見つかりません"):
            manager.delete_template("non_existent_user_template")


# ====================================
# update_template テスト
# ====================================
class TestUpdateTemplate:
    """update_template メソッドのテスト"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """テスト用の一時ディレクトリ"""
        builtin_dir = tmp_path / "templates" / "builtin"
        user_dir = tmp_path / "templates" / "user"
        builtin_dir.mkdir(parents=True, exist_ok=True)
        user_dir.mkdir(parents=True, exist_ok=True)

        # テスト用組み込みテンプレート
        builtin_template = {
            "id": "builtin_test",
            "label": "組み込みテスト",
            "description": "テスト用組み込みテンプレート",
            "category": "属性系",
            "attributes": ["属性1", "属性2"],
            "context": "組み込みコンテキスト",
            "batch_size": 5,
            "is_builtin": True,
            "created_at": "2026-02-17T00:00:00",
            "updated_at": "2026-02-17T00:00:00",
        }
        with open(builtin_dir / "builtin_test.json", "w", encoding="utf-8") as f:
            json.dump(builtin_template, f, ensure_ascii=False)

        return tmp_path

    def test_update_user_template(self, temp_dir):
        """userテンプレートの更新"""
        manager = TemplateManager(project_root=temp_dir)

        # ユーザーテンプレート作成
        user_template = InvestigationTemplate(
            id="user_update_test",
            label="更新前ラベル",
            category="カスタム",
            attributes=["属性A"],
            context="更新前コンテキスト",
        )
        manager.save_template(user_template)

        # 部分更新
        updated = manager.update_template(
            "user_update_test",
            label="更新後ラベル",
            context="更新後コンテキスト",
        )
        assert updated.label == "更新後ラベル"
        assert updated.context == "更新後コンテキスト"
        assert updated.attributes == ["属性A"]  # 変更していないフィールド

    def test_update_builtin_creates_copy(self, temp_dir):
        """builtinテンプレートの更新でコピーが作成される"""
        manager = TemplateManager(project_root=temp_dir)

        updated = manager.update_template(
            "builtin_test",
            label="カスタムラベル",
            attributes=["属性1", "属性2", "属性3"],
        )
        assert updated.id == "builtin_test_custom"
        assert updated.label == "カスタムラベル"
        assert updated.is_builtin is False
        assert len(updated.attributes) == 3

        # 元のbuiltinは変更されていないこと
        original = manager.get_template("builtin_test")
        assert original.is_builtin is True
        assert len(original.attributes) == 2

    def test_update_nonexistent_raises_key_error(self, temp_dir):
        """存在しないIDでKeyError"""
        manager = TemplateManager(project_root=temp_dir)

        with pytest.raises(KeyError):
            manager.update_template("nonexistent_template", label="新ラベル")


# ====================================
# export_to_json_bytes / import_from_json テスト
# ====================================
class TestExportImportJson:
    """JSON エクスポート/インポートのテスト"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """テスト用の一時ディレクトリ"""
        builtin_dir = tmp_path / "templates" / "builtin"
        user_dir = tmp_path / "templates" / "user"
        builtin_dir.mkdir(parents=True, exist_ok=True)
        user_dir.mkdir(parents=True, exist_ok=True)

        builtin_template = {
            "id": "export_test",
            "label": "エクスポートテスト",
            "description": "テスト用",
            "category": "属性系",
            "attributes": ["属性X", "属性Y"],
            "context": "テストコンテキスト",
            "batch_size": 5,
            "is_builtin": True,
            "created_at": "2026-02-17T00:00:00",
            "updated_at": "2026-02-17T00:00:00",
        }
        with open(builtin_dir / "export_test.json", "w", encoding="utf-8") as f:
            json.dump(builtin_template, f, ensure_ascii=False)

        return tmp_path

    def test_export_to_json_bytes_valid(self, temp_dir):
        """export_to_json_bytes: 有効なJSONバイト列が返ること"""
        manager = TemplateManager(project_root=temp_dir)
        result = manager.export_to_json_bytes("export_test")

        assert isinstance(result, bytes)
        data = json.loads(result.decode("utf-8"))
        assert data["id"] == "export_test"
        assert data["label"] == "エクスポートテスト"

    def test_export_to_json_bytes_is_builtin_false(self, temp_dir):
        """export_to_json_bytes: is_builtin=Falseであること"""
        manager = TemplateManager(project_root=temp_dir)
        result = manager.export_to_json_bytes("export_test")

        data = json.loads(result.decode("utf-8"))
        assert data["is_builtin"] is False

    def test_import_from_json_normal(self, temp_dir):
        """import_from_json: 正常インポート"""
        manager = TemplateManager(project_root=temp_dir)
        json_data = json.dumps({
            "id": "imported_test",
            "label": "インポートテスト",
            "category": "カスタム",
            "attributes": ["属性1", "属性2"],
            "context": "インポートコンテキスト",
        }, ensure_ascii=False)

        result = manager.import_from_json(json_data)
        assert result.id == "imported_test"
        assert result.label == "インポートテスト"
        assert result.is_builtin is False

        # 保存されていることを確認
        loaded = manager.get_template("imported_test")
        assert loaded.label == "インポートテスト"

    def test_import_from_json_invalid_json(self, temp_dir):
        """import_from_json: 不正JSONでValueError"""
        manager = TemplateManager(project_root=temp_dir)

        with pytest.raises(ValueError, match="不正なJSON"):
            manager.import_from_json("これはJSONではない{{{")

    def test_import_from_json_missing_required(self, temp_dir):
        """import_from_json: 必須フィールド不足でValueError"""
        manager = TemplateManager(project_root=temp_dir)
        json_data = json.dumps({
            "id": "missing_test",
            # label がない
            "attributes": ["A"],
        })

        with pytest.raises(ValueError, match="必須フィールドが不足"):
            manager.import_from_json(json_data)

    def test_import_from_json_disallowed_fields_excluded(self, temp_dir):
        """import_from_json: 不許可フィールドが除外されること"""
        manager = TemplateManager(project_root=temp_dir)
        json_data = json.dumps({
            "id": "sanitize_test",
            "label": "サニタイズテスト",
            "category": "カスタム",
            "attributes": ["属性1"],
            "malicious_field": "攻撃データ",
            "created_at": "2020-01-01T00:00:00",  # ホワイトリスト外
            "updated_at": "2020-01-01T00:00:00",  # ホワイトリスト外
        }, ensure_ascii=False)

        result = manager.import_from_json(json_data)
        assert result.id == "sanitize_test"
        # malicious_field は InvestigationTemplate に存在しない
        assert not hasattr(result, "malicious_field")

    def test_import_from_json_is_builtin_forced_false(self, temp_dir):
        """import_from_json: is_builtinが強制Falseであること"""
        manager = TemplateManager(project_root=temp_dir)
        json_data = json.dumps({
            "id": "force_false_test",
            "label": "強制Falseテスト",
            "category": "カスタム",
            "attributes": ["属性1"],
            "is_builtin": True,  # True を指定しても
        }, ensure_ascii=False)

        result = manager.import_from_json(json_data)
        assert result.is_builtin is False  # 強制的に False

    def test_roundtrip_export_import(self, temp_dir):
        """ラウンドトリップ: export_to_json_bytes → import_from_json"""
        manager = TemplateManager(project_root=temp_dir)

        # まずユーザーテンプレートを作成
        original = InvestigationTemplate(
            id="roundtrip_test",
            label="ラウンドトリップテスト",
            description="往復確認用",
            category="分類系",
            attributes=["属性A", "属性B", "属性C"],
            context="テストコンテキスト",
            batch_size=7,
            is_builtin=False,
        )
        manager.save_template(original)

        # エクスポート
        exported_bytes = manager.export_to_json_bytes("roundtrip_test")

        # 別IDで再インポート（同一IDだと上書きになるのでIDを変更）
        data = json.loads(exported_bytes.decode("utf-8"))
        data["id"] = "roundtrip_imported"
        json_bytes = json.dumps(data, ensure_ascii=False).encode("utf-8")
        imported = manager.import_from_json(json_bytes)

        assert imported.id == "roundtrip_imported"
        assert imported.label == original.label
        assert imported.description == original.description
        assert imported.category == original.category
        assert imported.attributes == original.attributes
        assert imported.context == original.context
        assert imported.is_builtin is False


# ====================================
# export_all_to_excel_bytes テスト
# ====================================
class TestExportAllToExcelBytes:
    """export_all_to_excel_bytes のテスト"""

    @pytest.fixture
    def temp_dir(self, tmp_path):
        """テスト用の一時ディレクトリ"""
        builtin_dir = tmp_path / "templates" / "builtin"
        user_dir = tmp_path / "templates" / "user"
        builtin_dir.mkdir(parents=True, exist_ok=True)
        user_dir.mkdir(parents=True, exist_ok=True)

        builtin_template = {
            "id": "excel_test",
            "label": "Excelテスト",
            "description": "テスト用",
            "category": "属性系",
            "attributes": ["属性X", "属性Y"],
            "context": "",
            "batch_size": 5,
            "is_builtin": True,
            "created_at": "2026-02-17T00:00:00",
            "updated_at": "2026-02-17T00:00:00",
        }
        with open(builtin_dir / "excel_test.json", "w", encoding="utf-8") as f:
            json.dump(builtin_template, f, ensure_ascii=False)

        return tmp_path

    def test_export_all_to_excel_bytes_valid(self, temp_dir):
        """export_all_to_excel_bytes: 有効なバイト列が返ること"""
        pytest.importorskip("openpyxl")
        manager = TemplateManager(project_root=temp_dir)
        result = manager.export_all_to_excel_bytes()

        assert isinstance(result, bytes)
        assert len(result) > 0

        # openpyxl で開けることを確認
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # ヘッダー行を確認
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=1, column=2).value == "名前"
        # データ行が存在すること
        assert ws.cell(row=2, column=1).value == "excel_test"
        wb.close()


# ====================================
# 実際の組み込みテンプレートとの整合性テスト
# ====================================
class TestBuiltinTemplates:
    """実際の組み込みテンプレートの整合性確認"""

    def test_builtin_templates_loadable(self):
        """組み込みテンプレートが全てロード可能か"""
        manager = TemplateManager()
        templates = manager.list_templates()

        # 少なくとも4つの組み込みテンプレートが存在するはず
        builtin_templates = [t for t in templates if t.is_builtin]
        assert len(builtin_templates) >= 4

    def test_video_streaming_template(self):
        """動画配信テンプレートの内容確認"""
        manager = TemplateManager()

        try:
            template = manager.get_template("動画配信_ジャンル")
            assert template.label == "動画配信サービス カテゴリ調査"
            assert template.category == "属性系"
            assert "邦画" in template.attributes
            assert "アニメ" in template.attributes
            assert template.is_builtin is True
        except KeyError:
            pytest.skip("動画配信_ジャンル テンプレートが存在しない")

    def test_area_8_template(self):
        """8地方区分テンプレートの内容確認"""
        manager = TemplateManager()

        try:
            template = manager.get_template("地方区分_8地方")
            assert template.label == "8地方区分"
            assert template.category == "地理系"
            assert len(template.attributes) == 8
            assert "北海道" in template.attributes
            assert "九州・沖縄" in template.attributes
            assert template.is_builtin is True
            # コンテキストが存在するか
            assert len(template.context) > 0
        except KeyError:
            pytest.skip("地方区分_8地方 テンプレートが存在しない")
