#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel読み書きハンドラー
======================
プレイヤーリストExcelの読み込み・書き出しを担当

対応フォーマット:
- クレジットカード プレイヤーリスト
- 動画配信サービス プレイヤーリスト
- 中古車販売店 プレイヤーリスト
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class PlayerData:
    """プレイヤーデータ"""
    row_index: int  # 元のExcel行番号
    player_name: str  # プレイヤー名（サービス名/企業名）
    official_url: str = ""  # 公式URL
    company_name: str = ""  # 運営会社名
    category: str = ""  # カテゴリ（業界固有）
    notes: str = ""  # 備考
    extra_data: dict = field(default_factory=dict)  # その他のカラムデータ


class ExcelHandler:
    """
    Excelファイルの読み書きを担当

    【対応列の自動検出】
    - プレイヤー名/サービス名/企業名 列を自動検出
    - URL/公式URL/公式サイト 列を自動検出
    - 運営会社/事業者 列を自動検出
    """

    # 列名の候補パターン（優先度順）
    PLAYER_NAME_PATTERNS = [
        r"^サービス名$", r"^プレイヤー名$", r"^ブランド名$", r"^企業名$", r"^会社名$",
        r"サービス名", r"プレイヤー名", r"ブランド名",
        r"player.*name", r"service.*name", r"brand",
    ]
    URL_PATTERNS = [
        r"^URL$", r"^公式URL$", r"^公式サイト$", r"^ホームページ$", r"^HP$",
        r"URL", r"公式URL", r"公式サイト", r"ホームページ", r"HP",
        r"official.*url", r"website", r"site",
    ]
    COMPANY_PATTERNS = [
        r"^事業者名$", r"^運営会社$", r"^運営元$",
        r"事業者名", r"事業者", r"運営会社", r"運営元",
        r"operator", r"company",
    ]

    # ヘッダー行検出用のキーワード（これらが含まれる行をヘッダーと判定）
    HEADER_KEYWORDS = [
        "サービス名", "プレイヤー名", "ブランド名", "企業名", "会社名",
        "事業者名", "調査票用No", "調査対象",
    ]

    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.header_row = 1
        self.column_map: dict[str, int] = {}  # 列名 -> 列インデックス

    def load(self, file_path: str | Path) -> list[PlayerData]:
        """
        Excelファイルを読み込み、プレイヤーデータのリストを返す

        Args:
            file_path: Excelファイルパス

        Returns:
            list[PlayerData]: プレイヤーデータのリスト
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            self.sheet = self.workbook.active

            # ヘッダー行を探す
            self._find_header_row()

            # 列マッピングを作成
            self._create_column_map()

            # データを読み込み
            players = []
            for row_idx in range(self.header_row + 1, self.sheet.max_row + 1):
                player = self._read_row(row_idx)
                if player and player.player_name.strip():
                    players.append(player)

            return players
        finally:
            self.workbook.close()

    def _find_header_row(self) -> None:
        """ヘッダー行を探す（キーワードマッチングで検出）"""
        best_row = 1
        best_score = 0

        for row_idx in range(1, min(15, self.sheet.max_row + 1)):
            row_values = [
                str(cell.value or "").strip()
                for cell in self.sheet[row_idx]
            ]
            row_text = " ".join(row_values)

            # ヘッダーキーワードのマッチ数をスコアとする
            score = 0
            for keyword in self.HEADER_KEYWORDS:
                if keyword in row_text:
                    score += 1

            # より多くのキーワードがマッチする行をヘッダーとする
            if score > best_score:
                best_score = score
                best_row = row_idx

        self.header_row = best_row

    def _create_column_map(self) -> None:
        """列名とインデックスのマッピングを作成"""
        self.column_map = {}

        header_cells = list(self.sheet[self.header_row])
        for col_idx, cell in enumerate(header_cells, 1):
            col_name = str(cell.value or "").strip()
            if not col_name:
                continue

            # 列名をそのまま保存
            self.column_map[col_name] = col_idx

            # 特殊な列を検出
            for pattern in self.PLAYER_NAME_PATTERNS:
                if re.search(pattern, col_name, re.IGNORECASE):
                    self.column_map["_player_name"] = col_idx
                    break

            for pattern in self.URL_PATTERNS:
                if re.search(pattern, col_name, re.IGNORECASE):
                    self.column_map["_url"] = col_idx
                    break

            for pattern in self.COMPANY_PATTERNS:
                if re.search(pattern, col_name, re.IGNORECASE):
                    self.column_map["_company"] = col_idx
                    break

    def _read_row(self, row_idx: int) -> Optional[PlayerData]:
        """1行をPlayerDataに変換"""
        logger = logging.getLogger(__name__)

        # プレイヤー名を取得
        player_name_col = self.column_map.get("_player_name")
        if player_name_col is None:
            # フォールバック: 列1を使用するが警告を出す
            player_name_col = 1
            if row_idx == self.header_row + 1:  # 最初の行でのみ警告
                logger.warning(
                    "プレイヤー名列が自動検出されませんでした。列1をフォールバックとして使用します。"
                )

        player_name = str(self.sheet.cell(row_idx, player_name_col).value or "").strip()

        if not player_name:
            return None

        # フォールバック列を使用している場合、数字のみの行はスキップ
        if "_player_name" not in self.column_map and player_name.isdigit():
            return None

        # URLを取得
        url_col = self.column_map.get("_url")
        official_url = ""
        if url_col:
            official_url = str(self.sheet.cell(row_idx, url_col).value or "").strip()

        # 運営会社を取得
        company_col = self.column_map.get("_company")
        company_name = ""
        if company_col:
            company_name = str(self.sheet.cell(row_idx, company_col).value or "").strip()

        # その他のデータを収集
        extra_data = {}
        for col_name, col_idx in self.column_map.items():
            if col_name.startswith("_"):
                continue
            value = self.sheet.cell(row_idx, col_idx).value
            if value is not None:
                extra_data[col_name] = str(value).strip()

        return PlayerData(
            row_index=row_idx,
            player_name=player_name,
            official_url=official_url,
            company_name=company_name,
            extra_data=extra_data,
        )

    def get_column_names(self) -> list[str]:
        """検出された列名のリストを返す"""
        return [k for k in self.column_map.keys() if not k.startswith("_")]


class ValidationReportExporter:
    """
    正誤チェック結果をExcelにエクスポート

    【出力形式】
    - アラートレベル別の色分け
    - 要確認行のハイライト
    """

    # アラートレベル別の色
    ALERT_COLORS = {
        "CRITICAL": "FF6B6B",   # 赤
        "WARNING": "FFD93D",    # 黄色
        "INFO": "6BCB77",       # 緑
        "OK": "FFFFFF",         # 白
        "UNCERTAIN": "FFA500",  # オレンジ
    }

    # ヘッダー列
    REPORT_COLUMNS = [
        "アラート",
        "プレイヤー名（元）",
        "プレイヤー名（現在）",
        "変更タイプ",
        "変更内容",
        "公式URL（元）",
        "公式URL（現在）",
        "運営会社（元）",
        "運営会社（現在）",
        "信頼度",
        "要確認フラグ",
        "関連ニュース",
        "情報ソース",
        "チェック日時",
    ]

    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "チェック結果"

    def export(
        self,
        results: list,  # list[ValidationResult]
        output_path: str | Path,
    ) -> Path:
        """
        チェック結果をExcelファイルに出力

        Args:
            results: ValidationResult のリスト
            output_path: 出力ファイルパス

        Returns:
            出力されたファイルのパス
        """
        output_path = Path(output_path)

        # ヘッダー行を作成
        self._write_header()

        # データ行を書き込み
        for row_idx, result in enumerate(results, start=2):
            self._write_row(row_idx, result)

        # 列幅を調整
        self._adjust_column_widths()

        # 保存
        self.workbook.save(output_path)
        return output_path

    def _write_header(self) -> None:
        """ヘッダー行を書き込み"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(self.REPORT_COLUMNS, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # ヘッダー行を固定
        self.sheet.freeze_panes = "A2"

    def _write_row(self, row_idx: int, result) -> None:
        """1行を書き込み"""
        # アラートレベルに応じた色
        alert_level = result.alert_level.name if hasattr(result.alert_level, "name") else str(result.alert_level)
        fill_color = self.ALERT_COLORS.get(alert_level, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # 要確認の場合はオレンジ背景
        if result.needs_manual_review:
            row_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # データを書き込み
        row_data = [
            result.alert_level.value if hasattr(result.alert_level, "value") else str(result.alert_level),
            result.player_name_original,
            result.player_name_current,
            result.change_type.value if hasattr(result.change_type, "value") else str(result.change_type),
            "\n".join(result.change_details) if result.change_details else "",
            result.url_original,
            result.url_current,
            getattr(result, "company_name_original", ""),
            getattr(result, "company_name_current", ""),
            f"{result.confidence * 100:.0f}%",
            "TRUE" if result.needs_manual_review else "FALSE",
            result.news_summary or "",
            "\n".join(result.source_urls) if result.source_urls else "",
            result.checked_at.strftime("%Y-%m-%d %H:%M:%S") if result.checked_at else "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = self.sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _adjust_column_widths(self) -> None:
        """列幅を調整"""
        column_widths = [8, 20, 20, 15, 40, 30, 30, 20, 20, 10, 12, 40, 40, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            self.sheet.column_dimensions[col_letter].width = width


class StoreInvestigationExporter:
    """
    店舗調査結果をExcelにエクスポート

    【出力形式】
    - 企業情報、店舗数
    - 調査モード、信頼度
    - 47都道府県別の店舗数
    - ソースURL（必須）
    """

    # 47都道府県リスト
    PREFECTURES = [
        "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
        "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
        "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
        "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
        "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
        "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
        "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県"
    ]

    # 信頼度別の色
    CONFIDENCE_COLORS = {
        "high": "6BCB77",     # 緑（0.8以上）
        "medium": "FFD93D",   # 黄色（0.5-0.8）
        "low": "FF6B6B",      # 赤（0.5未満）
    }

    # 基本ヘッダー列
    BASE_COLUMNS = [
        "企業名",
        "総店舗数",
        "直営店",
        "FC店",
        "調査モード",
        "信頼度",
        "要確認フラグ",
    ]

    # 後続ヘッダー列
    SUFFIX_COLUMNS = [
        "ソースURL",
        "備考",
        "調査日時",
    ]

    def __init__(self, include_prefectures: bool = True):
        """
        Args:
            include_prefectures: 都道府県別列を含めるか
        """
        self.include_prefectures = include_prefectures
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "店舗調査結果"

    def get_columns(self) -> list[str]:
        """出力列名のリストを取得"""
        columns = self.BASE_COLUMNS.copy()
        if self.include_prefectures:
            columns.extend(self.PREFECTURES)
        columns.extend(self.SUFFIX_COLUMNS)
        return columns

    def export(
        self,
        results: list,  # list[StoreInvestigationResult]
        output_path: str | Path,
    ) -> Path:
        """
        店舗調査結果をExcelファイルに出力

        Args:
            results: StoreInvestigationResult のリスト
            output_path: 出力ファイルパス

        Returns:
            出力されたファイルのパス
        """
        output_path = Path(output_path)

        # ヘッダー行を作成
        self._write_header()

        # データ行を書き込み
        for row_idx, result in enumerate(results, start=2):
            self._write_row(row_idx, result)

        # 列幅を調整
        self._adjust_column_widths()

        # 保存
        self.workbook.save(output_path)
        return output_path

    def _write_header(self) -> None:
        """ヘッダー行を書き込み"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        columns = self.get_columns()
        for col_idx, col_name in enumerate(columns, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # ヘッダー行を固定
        self.sheet.freeze_panes = "A2"

    def _write_row(self, row_idx: int, result) -> None:
        """1行を書き込み"""
        # 信頼度に応じた色
        if result.confidence >= 0.8:
            fill_color = self.CONFIDENCE_COLORS["high"]
        elif result.confidence >= 0.5:
            fill_color = self.CONFIDENCE_COLORS["medium"]
        else:
            fill_color = self.CONFIDENCE_COLORS["low"]

        # 要確認の場合は別色
        if result.needs_verification:
            row_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        else:
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # 基本データ
        row_data = [
            result.company_name,
            result.total_stores,
            result.direct_stores if result.direct_stores is not None else "",
            result.franchise_stores if result.franchise_stores is not None else "",
            result.investigation_mode,
            f"{result.confidence * 100:.0f}%",
            "TRUE" if result.needs_verification else "FALSE",
        ]

        # 都道府県別データ（○/×/?形式）
        if self.include_prefectures:
            pref_dist = result.prefecture_distribution or {}
            for pref in self.PREFECTURES:
                value = pref_dist.get(pref)
                if value is True:
                    row_data.append("○")  # 店舗あり
                elif value is False:
                    row_data.append("×")  # 店舗なし
                elif isinstance(value, int) and value > 0:
                    row_data.append("○")  # 旧形式（数値）の場合
                elif value == 0:
                    row_data.append("×")  # 0店舗
                else:
                    row_data.append("?")  # 不明

        # 後続データ
        source_urls = "\n".join(result.source_urls) if result.source_urls else ""
        investigation_date = (
            result.investigation_date.strftime("%Y-%m-%d %H:%M:%S")
            if result.investigation_date else ""
        )
        row_data.extend([
            source_urls,
            result.notes or "",
            investigation_date,
        ])

        # セルに書き込み
        for col_idx, value in enumerate(row_data, start=1):
            cell = self.sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _adjust_column_widths(self) -> None:
        """列幅を調整"""
        columns = self.get_columns()

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)

            # 列名に応じた幅を設定
            if col_name == "企業名":
                width = 25
            elif col_name in ("総店舗数", "直営店", "FC店"):
                width = 10
            elif col_name in ("調査モード", "信頼度", "要確認フラグ"):
                width = 12
            elif col_name == "ソースURL":
                width = 50
            elif col_name == "備考":
                width = 40
            elif col_name == "調査日時":
                width = 20
            elif col_name in self.PREFECTURES:
                width = 6  # 都道府県は狭めに
            else:
                width = 15

            self.sheet.column_dimensions[col_letter].width = width


class AttributeInvestigationExporter:
    """
    属性調査結果をExcelにエクスポート

    【出力形式】
    - 行: プレイヤー
    - 列: 属性 → ○/×/? マトリクス
    - 信頼度、ソースURL
    """

    # 信頼度別の色
    CONFIDENCE_COLORS = {
        "high": "6BCB77",     # 緑（0.8以上）
        "medium": "FFD93D",   # 黄色（0.5-0.8）
        "low": "FF6B6B",      # 赤（0.5未満）
    }

    # 属性値別の色
    ATTRIBUTE_COLORS = {
        True: "C6EFCE",   # 薄緑（○）
        False: "FFC7CE",  # 薄赤（×）
        None: "FFEB9C",   # 薄黄（?）
    }

    # 基本ヘッダー列
    BASE_COLUMNS = [
        "プレイヤー名",
    ]

    # 後続ヘッダー列
    SUFFIX_COLUMNS = [
        "信頼度",
        "要確認フラグ",
        "ソースURL",
        "調査日時",
    ]

    def __init__(self, attributes: list[str]):
        """
        Args:
            attributes: 属性名リスト（列として出力）
        """
        self.attributes = attributes
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "属性調査結果"

    def get_columns(self) -> list[str]:
        """出力列名のリストを取得"""
        columns = self.BASE_COLUMNS.copy()
        columns.extend(self.attributes)
        columns.extend(self.SUFFIX_COLUMNS)
        return columns

    def export(
        self,
        results: list,  # list[AttributeInvestigationResult]
        output_path: str | Path,
    ) -> Path:
        """
        属性調査結果をExcelファイルに出力

        Args:
            results: AttributeInvestigationResult のリスト
            output_path: 出力ファイルパス

        Returns:
            出力されたファイルのパス
        """
        output_path = Path(output_path)

        # ヘッダー行を作成
        self._write_header()

        # データ行を書き込み
        for row_idx, result in enumerate(results, start=2):
            self._write_row(row_idx, result)

        # 列幅を調整
        self._adjust_column_widths()

        # 保存
        self.workbook.save(output_path)
        return output_path

    def _write_header(self) -> None:
        """ヘッダー行を書き込み"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        columns = self.get_columns()
        for col_idx, col_name in enumerate(columns, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # ヘッダー行を固定
        self.sheet.freeze_panes = "A2"

    def _write_row(self, row_idx: int, result) -> None:
        """1行を書き込み"""
        # 信頼度に応じた行色
        if result.confidence >= 0.8:
            row_fill_color = self.CONFIDENCE_COLORS["high"]
        elif result.confidence >= 0.5:
            row_fill_color = self.CONFIDENCE_COLORS["medium"]
        else:
            row_fill_color = self.CONFIDENCE_COLORS["low"]

        col_idx = 1

        # プレイヤー名
        cell = self.sheet.cell(row=row_idx, column=col_idx, value=result.player_name)
        cell.alignment = Alignment(vertical="top")
        if result.needs_verification:
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        col_idx += 1

        # 属性マトリクス（○/×/?）
        attr_matrix = result.attribute_matrix or {}
        for attr in self.attributes:
            value = attr_matrix.get(attr)

            if value is True:
                display = "○"
                fill_color = self.ATTRIBUTE_COLORS[True]
            elif value is False:
                display = "×"
                fill_color = self.ATTRIBUTE_COLORS[False]
            else:
                display = "?"
                fill_color = self.ATTRIBUTE_COLORS[None]

            cell = self.sheet.cell(row=row_idx, column=col_idx, value=display)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        # 信頼度
        confidence_cell = self.sheet.cell(
            row=row_idx, column=col_idx,
            value=f"{result.confidence * 100:.0f}%"
        )
        confidence_cell.fill = PatternFill(
            start_color=row_fill_color, end_color=row_fill_color, fill_type="solid"
        )
        col_idx += 1

        # 要確認フラグ
        self.sheet.cell(
            row=row_idx, column=col_idx,
            value="TRUE" if result.needs_verification else "FALSE"
        )
        col_idx += 1

        # ソースURL
        source_urls = "\n".join(result.source_urls) if result.source_urls else ""
        cell = self.sheet.cell(row=row_idx, column=col_idx, value=source_urls)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        col_idx += 1

        # 調査日時
        investigation_date = (
            result.investigation_date.strftime("%Y-%m-%d %H:%M:%S")
            if result.investigation_date else ""
        )
        self.sheet.cell(row=row_idx, column=col_idx, value=investigation_date)

    def _adjust_column_widths(self) -> None:
        """列幅を調整"""
        columns = self.get_columns()

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)

            if col_name == "プレイヤー名":
                width = 25
            elif col_name in ("信頼度", "要確認フラグ"):
                width = 12
            elif col_name == "ソースURL":
                width = 50
            elif col_name == "調査日時":
                width = 20
            elif col_name in self.attributes:
                # 属性名の長さに応じて調整（最小6、最大15）
                width = max(6, min(15, len(col_name) * 2 + 2))
            else:
                width = 15

            self.sheet.column_dimensions[col_letter].width = width
