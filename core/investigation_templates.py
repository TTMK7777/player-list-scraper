#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
調査テンプレート管理
====================
属性調査のテンプレートを管理するモジュール。
組み込み（builtin）テンプレートとユーザー定義テンプレートをJSON形式で永続化。

【使用方法】
```python
from core.investigation_templates import TemplateManager

manager = TemplateManager()
templates = manager.list_templates(category="属性系")
template = manager.get_template("動画配信_ジャンル")
```
"""

import html
import io
import json
import logging
import os
import tempfile
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# プロジェクトルートとテンプレートディレクトリの絶対パス解決
PROJECT_ROOT = Path(__file__).parent.parent
BUILTIN_DIR = PROJECT_ROOT / "templates" / "builtin"
USER_DIR = PROJECT_ROOT / "templates" / "user"

# 有効なカテゴリ
VALID_CATEGORIES = ("属性系", "地理系", "分類系", "カスタム")


@dataclass
class InvestigationTemplate:
    """調査テンプレートのデータクラス。

    Attributes:
        id: スラッグ (例: "動画配信_ジャンル")
        label: 表示名
        description: 説明
        category: カテゴリ ("属性系" | "地理系" | "分類系" | "カスタム")
        attributes: 調査属性リスト
        context: 判定基準の説明 (LLMプロンプトに挿入)
        batch_size: バッチサイズ上書き (None=自動)
        is_builtin: 組み込みフラグ
        created_at: ISO datetime 文字列
        updated_at: ISO datetime 文字列
    """

    id: str
    label: str
    description: str = ""
    category: str = ""
    attributes: list[str] = field(default_factory=list)
    context: str = ""
    batch_size: Optional[int] = None
    is_builtin: bool = False
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())
    updated_at: str = field(default_factory=lambda: datetime.now().isoformat())

    def __post_init__(self):
        """バリデーション。"""
        if not self.id:
            raise ValueError("テンプレートIDは必須です")
        if not self.label:
            raise ValueError("テンプレートラベルは必須です")
        if self.category not in VALID_CATEGORIES:
            raise ValueError(
                f"無効なカテゴリ: '{self.category}' "
                f"(有効: {', '.join(VALID_CATEGORIES)})"
            )
        if not self.attributes:
            raise ValueError("属性リストは1つ以上必要です")

    def to_dict(self) -> dict:
        """テンプレートを辞書に変換する。

        Returns:
            JSON シリアライズ可能な辞書。
        """
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "InvestigationTemplate":
        """辞書からテンプレートを生成する。

        Args:
            data: テンプレートデータの辞書。

        Returns:
            InvestigationTemplate インスタンス。

        Raises:
            ValueError: 必須フィールドが不足している場合。
        """
        required_fields = ("id", "label", "category", "attributes")
        for f in required_fields:
            if f not in data:
                raise ValueError(f"必須フィールド '{f}' がありません")

        return cls(
            id=data["id"],
            label=data["label"],
            description=data.get("description", ""),
            category=data["category"],
            attributes=data["attributes"],
            context=data.get("context", ""),
            batch_size=data.get("batch_size"),
            is_builtin=data.get("is_builtin", False),
            created_at=data.get("created_at", datetime.now().isoformat()),
            updated_at=data.get("updated_at", datetime.now().isoformat()),
        )


class TemplateManager:
    """テンプレートの CRUD 操作を提供するマネージャークラス。

    builtin ディレクトリと user ディレクトリの両方からテンプレートを読み込み、
    ユーザーテンプレートの保存・削除を行う。

    Attributes:
        builtin_dir: 組み込みテンプレートディレクトリ。
        user_dir: ユーザーテンプレートディレクトリ。
    """

    def __init__(self, project_root: Optional[Path] = None):
        """TemplateManager を初期化する。

        Args:
            project_root: プロジェクトルートパス。None の場合はデフォルトを使用。
        """
        if project_root is not None:
            self.builtin_dir = project_root / "templates" / "builtin"
            self.user_dir = project_root / "templates" / "user"
        else:
            self.builtin_dir = BUILTIN_DIR
            self.user_dir = USER_DIR

        # ユーザーディレクトリを自動作成
        self.user_dir.mkdir(parents=True, exist_ok=True)

    def _load_from_dir(self, directory: Path) -> list[InvestigationTemplate]:
        """指定ディレクトリ内の全 JSON ファイルを読み込む。

        Args:
            directory: 読み込み対象ディレクトリ。

        Returns:
            テンプレートのリスト。読み込みエラーのファイルはスキップ。
        """
        templates = []
        if not directory.exists():
            return templates

        for json_file in sorted(directory.glob("*.json")):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                templates.append(InvestigationTemplate.from_dict(data))
            except (json.JSONDecodeError, ValueError, KeyError) as e:
                # 壊れたファイルはスキップ
                logger.warning(f"テンプレート読み込みエラー: {json_file} - {e}")
                continue

        return templates

    def list_templates(
        self, category: Optional[str] = None
    ) -> list[InvestigationTemplate]:
        """全テンプレートを取得する。

        builtin と user の両方を返す。category を指定するとフィルタリング。

        Args:
            category: フィルタするカテゴリ。None の場合は全件返却。

        Returns:
            テンプレートのリスト。
        """
        templates = self._load_from_dir(self.builtin_dir)
        templates.extend(self._load_from_dir(self.user_dir))

        if category is not None:
            templates = [t for t in templates if t.category == category]

        return templates

    def get_template(self, template_id: str) -> InvestigationTemplate:
        """ID でテンプレートを検索する。

        builtin -> user の順で検索し、最初に見つかったものを返す。

        Args:
            template_id: テンプレートID。

        Returns:
            テンプレートインスタンス。

        Raises:
            KeyError: テンプレートが見つからない場合。
        """
        # builtin を先に検索
        for directory in (self.builtin_dir, self.user_dir):
            json_path = directory / f"{template_id}.json"
            if json_path.exists():
                try:
                    with open(json_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    return InvestigationTemplate.from_dict(data)
                except (json.JSONDecodeError, ValueError, KeyError):
                    continue

        raise KeyError(f"テンプレート '{template_id}' が見つかりません")

    def save_template(self, template: InvestigationTemplate) -> Path:
        """ユーザーテンプレートを保存する。

        Args:
            template: 保存するテンプレート。

        Returns:
            保存先のファイルパス。

        Raises:
            PermissionError: 組み込みテンプレートを保存しようとした場合。
        """
        if template.is_builtin:
            raise PermissionError(
                "組み込みテンプレートは保存できません。"
                "is_builtin=False に変更してください。"
            )

        template.updated_at = datetime.now().isoformat()
        save_path = self.user_dir / f"{template.id}.json"

        fd, tmp_path = tempfile.mkstemp(dir=str(save_path.parent), suffix='.tmp')
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                json.dump(template.to_dict(), f, ensure_ascii=False, indent=4)
            os.replace(tmp_path, str(save_path))
        except BaseException:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise

        return save_path

    def delete_template(self, template_id: str) -> bool:
        """ユーザーテンプレートを削除する。

        Args:
            template_id: 削除するテンプレートのID。

        Returns:
            削除成功なら True。

        Raises:
            PermissionError: 組み込みテンプレートを削除しようとした場合。
            KeyError: テンプレートが見つからない場合。
        """
        # 組み込みテンプレートの削除を拒否
        builtin_path = self.builtin_dir / f"{template_id}.json"
        if builtin_path.exists():
            raise PermissionError(
                f"組み込みテンプレート '{template_id}' は削除できません"
            )

        user_path = self.user_dir / f"{template_id}.json"
        if not user_path.exists():
            raise KeyError(f"テンプレート '{template_id}' が見つかりません")

        user_path.unlink()
        return True

    def import_from_text(
        self, text: str, separator: str = ","
    ) -> list[str]:
        """テキストから属性リストを抽出する。

        カンマ・改行・指定セパレータで分割し、空白をトリミング。

        Args:
            text: 属性が含まれたテキスト。
            separator: 区切り文字（デフォルト: カンマ）。

        Returns:
            属性文字列のリスト（空文字列は除外）。
        """
        # 改行で分割してからセパレータで分割
        items = []
        for line in text.splitlines():
            for item in line.split(separator):
                stripped = item.strip()
                if stripped:
                    items.append(stripped)
        return items

    def import_from_excel(
        self, file_path: Path, column: int = 0
    ) -> list[str]:
        """Excel ファイルから属性リストを読み取る。

        Args:
            file_path: Excel ファイルのパス。
            column: 読み取る列のインデックス（0始まり）。

        Returns:
            属性文字列のリスト（空セルは除外）。

        Raises:
            FileNotFoundError: ファイルが存在しない場合。
            ImportError: openpyxl がインストールされていない場合。
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")

        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Excel読み込みには openpyxl が必要です。"
                "pip install openpyxl でインストールしてください。"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        items = []
        for row in ws.iter_rows(min_col=column + 1, max_col=column + 1):
            cell = row[0]
            if cell.value is not None:
                value = str(cell.value).strip()
                if value:
                    items.append(value)

        wb.close()
        return items

    def export_template(self, template_id: str) -> dict:
        """テンプレートを辞書として返す。

        Args:
            template_id: エクスポートするテンプレートのID。

        Returns:
            JSON シリアライズ可能な辞書。

        Raises:
            KeyError: テンプレートが見つからない場合。
        """
        template = self.get_template(template_id)
        return template.to_dict()

    def update_template(
        self,
        template_id: str,
        label: str | None = None,
        description: str | None = None,
        attributes: list[str] | None = None,
        context: str | None = None,
        batch_size: int | None = None,
    ) -> InvestigationTemplate:
        """既存テンプレートを部分更新する。

        Noneのフィールドは変更しない。builtin テンプレートの場合は
        コピーとしてuser保存する（IDに ``_custom`` 付与、ラベルに
        「（カスタム）」付与）。

        Args:
            template_id: 更新対象のテンプレートID。
            label: 新しいラベル（Noneなら変更なし）。
            description: 新しい説明（Noneなら変更なし）。
            attributes: 新しい属性リスト（Noneなら変更なし）。
            context: 新しいコンテキスト（Noneなら変更なし）。
            batch_size: 新しいバッチサイズ（Noneなら変更なし）。

        Returns:
            更新（または新規作成）された InvestigationTemplate。

        Raises:
            KeyError: テンプレートが見つからない場合。
        """
        template = self.get_template(template_id)

        if template.is_builtin:
            # builtin の場合はコピーとして保存
            new_id = f"{template_id}_custom"
            new_label = label if label is not None else f"{template.label}（カスタム）"
            template = InvestigationTemplate(
                id=new_id,
                label=new_label,
                description=description if description is not None else template.description,
                category=template.category,
                attributes=attributes if attributes is not None else list(template.attributes),
                context=context if context is not None else template.context,
                batch_size=batch_size if batch_size is not None else template.batch_size,
                is_builtin=False,
            )
        else:
            # user テンプレートは直接更新
            if label is not None:
                template.label = label
            if description is not None:
                template.description = description
            if attributes is not None:
                template.attributes = attributes
            if context is not None:
                template.context = context
            if batch_size is not None:
                template.batch_size = batch_size
            template.is_builtin = False

        self.save_template(template)
        return template

    def export_to_json_bytes(self, template_id: str) -> bytes:
        """テンプレートをJSON bytes として返す。

        ``export_template()`` の結果を JSON bytes 化する。
        ``is_builtin`` は False にリセットされる。

        Args:
            template_id: エクスポートするテンプレートのID。

        Returns:
            JSON エンコードされた bytes。

        Raises:
            KeyError: テンプレートが見つからない場合。
        """
        data = self.export_template(template_id)
        data["is_builtin"] = False
        return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")

    def import_from_json(self, json_data: bytes | str) -> InvestigationTemplate:
        """JSON データからテンプレートをインポートする。

        許可フィールドのホワイトリストで sanitize し、
        ``is_builtin`` は強制的に False にする。
        各属性に ``html.escape()`` を適用して XSS ペイロードを防止する。

        Args:
            json_data: JSON バイト列または文字列。

        Returns:
            インポートされた InvestigationTemplate。

        Raises:
            ValueError: 不正な JSON または必須フィールド不足の場合。
        """
        if isinstance(json_data, bytes):
            json_data = json_data.decode("utf-8")

        try:
            raw = json.loads(json_data)
        except (json.JSONDecodeError, TypeError) as e:
            raise ValueError(f"不正なJSON: {e}")

        if not isinstance(raw, dict):
            raise ValueError("JSONはオブジェクト形式でなければなりません")

        # 必須フィールドチェック
        required = ("id", "label", "attributes")
        missing = [f for f in required if f not in raw]
        if missing:
            raise ValueError(f"必須フィールドが不足しています: {', '.join(missing)}")

        # ホワイトリストで sanitize
        allowed_fields = {"id", "label", "description", "category", "attributes", "context", "batch_size"}
        sanitized = {k: v for k, v in raw.items() if k in allowed_fields}

        # IDのパストラバーサル防止（ファイル名に使用されるため）
        if "id" in sanitized:
            import re
            safe_id = re.sub(r"[^\w\u3000-\u9FFF\uF900-\uFAFF]", "_", str(sanitized["id"]))
            safe_id = re.sub(r"_+", "_", safe_id).strip("_")
            sanitized["id"] = safe_id or "imported_template"

        # category が未指定の場合はデフォルト値を設定
        if "category" not in sanitized:
            sanitized["category"] = "カスタム"

        # is_builtin は強制 False
        sanitized["is_builtin"] = False

        # attributes の各要素に html.escape 適用
        if "attributes" in sanitized and isinstance(sanitized["attributes"], list):
            sanitized["attributes"] = [html.escape(str(a)) for a in sanitized["attributes"]]

        template = InvestigationTemplate.from_dict(sanitized)
        self.save_template(template)
        return template

    def export_all_to_excel_bytes(self, category: str | None = None) -> bytes:
        """テンプレート一覧を Excel bytes として返す。

        列: ID / 名前 / カテゴリ / 属性数 / 属性リスト（カンマ区切り）/
        判定基準 / タイプ(builtin/user) / 日時

        Args:
            category: フィルタするカテゴリ。None の場合は全件。

        Returns:
            Excel ファイルの bytes データ。
        """
        import openpyxl

        templates = self.list_templates(category=category)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "テンプレート一覧"

        # ヘッダー
        headers = ["ID", "名前", "カテゴリ", "属性数", "属性リスト", "判定基準", "タイプ", "更新日時"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # データ行
        for row_idx, tmpl in enumerate(templates, 2):
            ws.cell(row=row_idx, column=1, value=tmpl.id)
            ws.cell(row=row_idx, column=2, value=tmpl.label)
            ws.cell(row=row_idx, column=3, value=tmpl.category)
            ws.cell(row=row_idx, column=4, value=len(tmpl.attributes))
            ws.cell(row=row_idx, column=5, value=", ".join(tmpl.attributes))
            ws.cell(row=row_idx, column=6, value=tmpl.context)
            ws.cell(row=row_idx, column=7, value="builtin" if tmpl.is_builtin else "user")
            ws.cell(row=row_idx, column=8, value=tmpl.updated_at)

        buffer = io.BytesIO()
        wb.save(buffer)
        wb.close()
        return buffer.getvalue()
