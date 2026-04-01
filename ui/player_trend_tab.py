#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
プレイヤーの最新動向 UIタブ（統合）
====================================
正誤チェック + 新規参入検出 + 最新版リスト作成を1タブに統合。
3つのサブタブ（st.tabs）で構成。
"""

import html
import io
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st

from core.async_helpers import run_async
from core.excel_handler import ExcelHandler, PlayerData
from core.llm_client import LLMClient
from investigators.base import (
    AlertLevel,
    ChangeType,
    NewcomerCandidate,
    ValidationResult,
)
from investigators.newcomer_detector import NewcomerDetector
from investigators.player_validator import PlayerValidator
from ui.common import (
    display_actual_cost,
    display_cost_estimate,
    display_filter_multiselect,
    display_progress_log,
    display_verification_badge,
    number_input_with_max,
    select_sheet_if_multiple,
)


# ====================================
# セッション状態初期化
# ====================================
def _init_session_state() -> None:
    """trend_ プレフィックス付きセッション状態を初期化"""
    defaults = {
        "trend_players": [],
        "trend_val_results": [],
        "trend_val_is_running": False,
        "trend_newcomer_candidates": [],
        "trend_newcomer_is_running": False,
    }
    for key, default_value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_value


# ====================================
# 共有 Excel アップロード
# ====================================
def _render_shared_upload() -> None:
    """サブタブ共通の Excel アップロード UI"""

    st.subheader("📂 Excelアップロード")
    st.info(
        "既存プレイヤーリストをアップロードしてください。"
        "3つのサブタブ（変更点調査・新規参入検出・最新版リスト作成）で共有されます。"
    )

    uploaded_file = st.file_uploader(
        "プレイヤーリストExcelをアップロード",
        type=["xlsx", "xls"],
        help="サービス名/プレイヤー名、公式URL を含むExcelファイル",
        key="trend_excel_upload",
    )

    if uploaded_file:
        try:
            temp_dir = Path(tempfile.gettempdir()) / "player_trend"
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_path = temp_dir / Path(uploaded_file.name).name
            temp_path.write_bytes(uploaded_file.getvalue())

            selected_sheet = select_sheet_if_multiple(temp_path, "trend")
            handler = ExcelHandler()
            players = handler.load_multiple(temp_path, sheet_names=selected_sheet)
            st.session_state.trend_players = players

            st.success(f"✅ {len(players)}件のプレイヤーを読み込みました")

            # フォールバック警告の表示
            if handler.warnings:
                for warn_msg in handler.warnings:
                    st.warning(f"⚠️ {warn_msg}")

            with st.expander("📋 検出された列"):
                cols = handler.get_column_names()
                st.write(cols)

            with st.expander("👀 データプレビュー（先頭10件）"):
                preview_data = []
                for p in players[:10]:
                    preview_data.append({
                        "プレイヤー名": p.player_name,
                        "公式URL": (
                            p.official_url[:50] + "..."
                            if len(p.official_url) > 50
                            else p.official_url
                        ),
                        "運営会社": p.company_name,
                    })
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)

        except Exception as e:
            st.error(f"❌ Excelの読み込みに失敗: {e}")
            st.session_state.trend_players = []


# ====================================
# サブタブ 1: 変更点調査
# ====================================
async def _run_validation(
    players: list[PlayerData],
    industry: str,
    definition: str,
    progress_container,
    status_container,
) -> list[ValidationResult]:
    """正誤チェックを実行"""

    logs: list[str] = []

    def on_progress(current: int, total: int, name: str) -> None:
        log_msg = f"[{current}/{total}] チェック中: {name}"
        logs.append(log_msg)
        display_progress_log(logs, progress_container)

    status_container.info(f"🔍 {len(players)}件のプレイヤーをチェック中...")

    try:
        llm = LLMClient()
        validator = PlayerValidator(llm_client=llm)

        results = await validator.validate_batch(
            players,
            industry=industry,
            definition=definition,
            on_progress=on_progress,
            delay_seconds=1.5,
        )

        status_container.success(f"✅ チェック完了: {len(results)}件")
        st.session_state.trend_val_last_llm = llm
        return results

    except Exception as e:
        status_container.error(f"❌ エラー: {type(e).__name__}: {str(e)}")
        return []


def _display_validation_summary(results: list[ValidationResult]) -> None:
    """正誤チェック結果サマリーを表示"""

    alert_counts = {
        AlertLevel.CRITICAL: 0,
        AlertLevel.WARNING: 0,
        AlertLevel.INFO: 0,
        AlertLevel.OK: 0,
    }
    uncertain_count = 0

    for result in results:
        alert_counts[result.alert_level] = alert_counts.get(result.alert_level, 0) + 1
        if result.needs_manual_review:
            uncertain_count += 1

    st.markdown("### 📊 チェック結果サマリー")

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("🔴 緊急（撤退・統合）", f"{alert_counts[AlertLevel.CRITICAL]}件")
    with col2:
        st.metric("🟡 警告（名称変更）", f"{alert_counts[AlertLevel.WARNING]}件")
    with col3:
        st.metric("🟢 情報（URL変更等）", f"{alert_counts[AlertLevel.INFO]}件")
    with col4:
        st.metric("✅ 変更なし", f"{alert_counts[AlertLevel.OK]}件")
    with col5:
        st.metric("⚠️ 要確認", f"{uncertain_count}件")


def _display_validation_table(results: list[ValidationResult]) -> None:
    """正誤チェック結果テーブルをフィルター付きで表示"""

    alert_labels = [level.value for level in AlertLevel]
    col_filter1, col_filter2 = st.columns(2)

    with col_filter1:
        selected_alerts = display_filter_multiselect(
            "アラートレベルで絞り込み",
            options=alert_labels,
            key="trend_val_filter_alert",
        )

    with col_filter2:
        show_attention_only = st.checkbox(
            "対応が必要なもののみ表示",
            value=False,
            key="trend_val_filter_manual",
        )

    def _needs_attention(r: ValidationResult) -> bool:
        """CRITICAL/WARNING/要確認のいずれかに該当"""
        return (
            r.alert_level in (AlertLevel.CRITICAL, AlertLevel.WARNING)
            or r.needs_manual_review
        )

    filtered_results = [
        r for r in results
        if r.alert_level.value in selected_alerts
        and (not show_attention_only or _needs_attention(r))
    ]

    st.caption(f"表示中: {len(filtered_results)} / {len(results)} 件")

    alert_order = {
        AlertLevel.CRITICAL: 0,
        AlertLevel.WARNING: 1,
        AlertLevel.INFO: 2,
        AlertLevel.OK: 3,
    }
    sorted_results = sorted(
        filtered_results,
        key=lambda r: (alert_order.get(r.alert_level, 4), not r.needs_manual_review),
    )

    data = []
    for result in sorted_results:
        data.append({
            "アラート": result.alert_level.value,
            "プレイヤー名（元）": result.player_name_original,
            "プレイヤー名（現在）": result.player_name_current,
            "変更タイプ": result.change_type.value,
            "変更内容": " / ".join(result.change_details) if result.change_details else "-",
            "公式URL": result.url_current or result.url_original or "",
            "情報ソース": result.source_urls[0] if result.source_urls else "",
            "要確認": "⚠️" if result.needs_manual_review else "",
        })

    df = pd.DataFrame(data)

    st.dataframe(
        df,
        use_container_width=True,
        height=400,
        column_config={
            "アラート": st.column_config.TextColumn("アラート", width="small"),
            "プレイヤー名（元）": st.column_config.TextColumn("プレイヤー名（元）", width="medium"),
            "プレイヤー名（現在）": st.column_config.TextColumn("プレイヤー名（現在）", width="medium"),
            "変更タイプ": st.column_config.TextColumn("変更タイプ", width="small"),
            "変更内容": st.column_config.TextColumn("変更内容", width="large"),
            "公式URL": st.column_config.LinkColumn("公式URL", width="medium", display_text="開く"),
            "情報ソース": st.column_config.LinkColumn("情報ソース", width="medium", display_text="開く"),
            "要確認": st.column_config.TextColumn("要確認", width="small"),
        },
    )


def _export_validation_results(results: list[ValidationResult]) -> bytes:
    """正誤チェック結果をExcelエクスポート"""

    report_data = []
    for result in results:
        report_data.append({
            "アラート": result.alert_level.value,
            "プレイヤー名（元）": result.player_name_original,
            "プレイヤー名（現在）": result.player_name_current,
            "変更タイプ": result.change_type.value,
            "変更内容": "\n".join(result.change_details) if result.change_details else "",
            "公式URL（元）": result.url_original,
            "公式URL（現在）": result.url_current,
            "運営会社（元）": result.company_name_original,
            "運営会社（現在）": result.company_name_current,
            "要確認フラグ": "TRUE" if result.needs_manual_review else "FALSE",
            "関連ニュース": result.news_summary,
            "情報ソース": "\n".join(result.source_urls) if result.source_urls else "",
            "チェック日時": (
                result.checked_at.strftime("%Y-%m-%d %H:%M:%S") if result.checked_at else ""
            ),
        })

    df_report = pd.DataFrame(report_data)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_report.to_excel(writer, index=False, sheet_name="チェック結果")

        from openpyxl.utils import get_column_letter

        worksheet = writer.sheets["チェック結果"]
        for idx, col in enumerate(df_report.columns):
            max_length = max(
                df_report[col].astype(str).str.len().max(),
                len(col),
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)

    return buffer.getvalue()


def _render_validation_subtab(industry: str, definition: str) -> None:
    """サブタブ 1: 変更点調査"""

    st.markdown(
        "既存プレイヤーリストの**撤退・統合・名称変更**をGemini AIが自動検出します。"
        "アラートレベル別にExcelレポートを出力できます。"
    )

    players = st.session_state.trend_players

    if not players:
        st.warning("先にExcelファイルをアップロードしてください。")
        return

    # チェック件数入力
    check_limit = number_input_with_max(
        label="チェック件数",
        max_value=len(players),
        default_value=min(10, len(players)),
        key="trend_val_check_limit",
        help="APIコスト削減のため、最初は少数でテストしてください",
    )

    # コスト概算表示
    cost = PlayerValidator.estimate_cost(check_limit)
    display_cost_estimate(
        call_count=cost["call_count"],
        cost_per_call=cost["cost_per_call"],
        label="正誤チェック",
    )

    # 実行ボタン
    run_button = st.button(
        "🚀 変更点調査を開始",
        type="primary",
        disabled=st.session_state.trend_val_is_running,
        use_container_width=True,
        key="trend_val_run_button",
    )

    st.divider()

    if run_button:
        st.session_state.trend_val_is_running = True

        progress_container = st.empty()
        status_container = st.empty()

        players_to_check = players[:check_limit]
        # UI境界: 空文字→None正規化
        industry_normalized = (industry.strip() or None) if industry else None

        try:
            results = run_async(_run_validation(
                players_to_check,
                industry=industry_normalized,
                definition=definition,
                progress_container=progress_container,
                status_container=status_container,
            ))

            st.session_state.trend_val_results = results
        except Exception as e:
            st.error(f"予期しないエラーが発生しました: {type(e).__name__}: {e}")
        finally:
            st.session_state.trend_val_is_running = False

    # 結果表示
    if st.session_state.trend_val_results:
        results = st.session_state.trend_val_results

        # 実コスト表示
        if "trend_val_last_llm" in st.session_state:
            display_actual_cost(st.session_state.trend_val_last_llm)

        _display_validation_summary(results)

        st.divider()

        st.subheader("📋 詳細結果（アラートレベル順）")
        _display_validation_table(results)

        st.divider()

        st.subheader("📥 結果エクスポート")

        col1, col2 = st.columns(2)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        with col1:
            excel_data = _export_validation_results(results)
            st.download_button(
                "📥 Excel ダウンロード（チェック結果）",
                excel_data,
                f"trend_validation_{timestamp}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="trend_val_export_excel",
            )

        with col2:
            csv_data = [result.to_dict() for result in results]
            df_csv = pd.DataFrame(csv_data)
            csv_bytes = df_csv.to_csv(index=False).encode("utf-8-sig")

            st.download_button(
                "📥 CSV ダウンロード",
                csv_bytes,
                f"trend_validation_{timestamp}.csv",
                "text/csv",
                use_container_width=True,
                key="trend_val_export_csv",
            )


# ====================================
# サブタブ 2: 新規参入検出
# ====================================
def _render_newcomer_subtab(industry: str, definition: str) -> None:
    """サブタブ 2: 新規参入検出"""

    st.markdown(
        "既存プレイヤーリストにない**新規参入企業**をAIが自動検索します。"
        "アップロード済みのリストから既存プレイヤーを取得します。"
    )

    players = st.session_state.trend_players

    if not players:
        st.warning("先にExcelファイルをアップロードしてください。")
        return

    existing_names = [p.player_name for p in players]
    st.info(f"📋 既存プレイヤー: {len(existing_names)}件（アップロードファイルから取得）")

    # コスト概算表示
    if industry:
        cost = NewcomerDetector.estimate_cost()
        display_cost_estimate(
            call_count=cost["call_count"],
            cost_per_call=cost["cost_per_call"],
            label="新規参入検出",
        )

    # 実行ボタン
    run_button = st.button(
        "🆕 新規参入を検索",
        type="primary",
        disabled=(
            not industry
            or st.session_state.trend_newcomer_is_running
        ),
        use_container_width=True,
        key="trend_newcomer_run_button",
    )

    if not industry:
        st.warning("業界が入力されていません（メイン画面で設定してください）。")

    st.divider()

    if run_button:
        st.session_state.trend_newcomer_is_running = True

        progress_container = st.empty()
        status_container = st.empty()

        def on_progress(current: int, total: int, name: str) -> None:
            safe_name = html.escape(name)
            progress_container.markdown(
                f'<div class="progress-log">[{current}/{total}] {safe_name}</div>',
                unsafe_allow_html=True,
            )

        # UI境界: 空文字→None正規化
        industry_normalized = (industry.strip() or None) if industry else None
        status_container.info("🔍 新規参入候補を検索中...")

        try:
            llm = LLMClient()
            detector = NewcomerDetector(llm_client=llm)

            candidates = run_async(detector.detect(
                industry=industry_normalized,
                existing_players=existing_names,
                definition=definition,
                on_progress=on_progress,
            ))

            st.session_state.trend_newcomer_candidates = candidates
            status_container.success(f"✅ 検出完了: {len(candidates)}件の候補")

        except Exception as e:
            status_container.error(f"❌ エラー: {type(e).__name__}: {str(e)}")
            st.session_state.trend_newcomer_candidates = []
        finally:
            st.session_state.trend_newcomer_is_running = False

    # 結果表示
    if st.session_state.trend_newcomer_candidates:
        candidates = st.session_state.trend_newcomer_candidates

        st.subheader("🆕 候補一覧")

        st.warning(
            "候補はAIの提案です。必ず手動確認してからエクスポートしてください。\n\n"
            "URL検証バッジの見方:\n"
            "- URL検証済み: 公式サイトの存在を確認\n"
            "- URL不明: URLにアクセスできなかった（要注意）\n"
            "- 未検証: URLが提供されていない"
        )

        for i, candidate in enumerate(candidates):
            badge = display_verification_badge(candidate.verification_status)

            with st.container():
                col1, col2, col3 = st.columns([0.5, 3, 3])

                with col1:
                    key = f"trend_newcomer_select_{i}"
                    st.session_state.setdefault(key, candidate.verification_status == "verified")
                    st.checkbox("", key=key, label_visibility="collapsed")

                with col2:
                    st.markdown(
                        f"**{html.escape(candidate.player_name)}** {badge}",
                        unsafe_allow_html=True,
                    )
                    if candidate.official_url:
                        st.caption(candidate.official_url)

                with col3:
                    st.caption(
                        f"運営: {candidate.company_name}" if candidate.company_name else ""
                    )
                    st.caption(
                        f"理由: {candidate.reason}" if candidate.reason else ""
                    )

            st.divider()

        # 選択状況サマリー
        selected_count = sum(
            1 for i in range(len(candidates))
            if st.session_state.get(f"trend_newcomer_select_{i}", False)
        )
        st.info(f"選択中: {selected_count}件 / {len(candidates)}件")


# ====================================
# サブタブ 3: 最新版リスト作成
# ====================================
def _compile_final_list(
    players: list[PlayerData],
    val_results: list[ValidationResult],
    newcomer_candidates: list[NewcomerCandidate],
) -> pd.DataFrame:
    """正誤チェック結果と新規参入候補からリストをコンパイル

    コンパイルロジック:
    - 正誤チェック結果から:
      - WITHDRAWAL / MERGER → 除外
      - SERVICE_RENAME → player_name_current を新名称として採用
      - COMPANY_RENAME → company_name_current を更新
      - URL_CHANGE → url_current を更新
      - その他（OK, INFO等） → そのまま保持
    - 新規参入候補（チェックボックス選択済みのみ）→ 新規追加

    Args:
        players: 元のプレイヤーリスト
        val_results: 正誤チェック結果
        newcomer_candidates: 新規参入候補（選択済みのみ）

    Returns:
        コンパイル済みの DataFrame
    """
    rows = []

    # 正誤チェック結果をプレイヤー名でインデックス化
    result_map: dict[str, ValidationResult] = {}
    for result in val_results:
        result_map[result.player_name_original] = result

    # 元リストのプレイヤーを処理
    for player in players:
        result = result_map.get(player.player_name)

        if result is not None:
            # チェック済みプレイヤー
            if result.change_type in (ChangeType.WITHDRAWAL, ChangeType.MERGER):
                # 撤退・統合 → 除外
                continue

            # 名称・URL・会社名の更新
            player_name = player.player_name
            official_url = player.official_url
            company_name = player.company_name
            source = "元リスト"
            remarks = ""

            if result.change_type == ChangeType.SERVICE_RENAME:
                player_name = result.player_name_current
                source = "名称変更"
                remarks = f"{result.player_name_original} → {result.player_name_current}"

            if result.change_type == ChangeType.COMPANY_RENAME:
                company_name = result.company_name_current or company_name
                if not remarks:
                    remarks = f"会社名変更: {result.company_name_original} → {result.company_name_current}"

            if result.change_type == ChangeType.URL_CHANGE:
                official_url = result.url_current or official_url
                if not remarks:
                    remarks = f"URL変更: {result.url_original} → {result.url_current}"

            # その他の変更詳細があれば備考に追記
            if result.change_details and not remarks:
                remarks = " / ".join(result.change_details)

            rows.append({
                "含む": True,
                "プレイヤー名": player_name,
                "公式URL": official_url,
                "運営会社": company_name,
                "ソース": source,
                "備考": remarks,
            })
        else:
            # 未チェックプレイヤー → そのまま保持
            rows.append({
                "含む": True,
                "プレイヤー名": player.player_name,
                "公式URL": player.official_url,
                "運営会社": player.company_name,
                "ソース": "元リスト",
                "備考": "",
            })

    # 選択済み新規参入候補を追加
    for candidate in newcomer_candidates:
        rows.append({
            "含む": True,
            "プレイヤー名": candidate.player_name,
            "公式URL": candidate.official_url,
            "運営会社": candidate.company_name,
            "ソース": "新規参入",
            "備考": candidate.reason or "",
        })

    return pd.DataFrame(rows)


def _export_compiled_excel(df: pd.DataFrame) -> bytes:
    """コンパイル済みリストを Excel にエクスポート

    Args:
        df: 「含む」列が True の行のみを出力対象とする DataFrame

    Returns:
        Excel ファイルのバイト列
    """
    # 「含む」が True の行のみ
    df_export = df[df["含む"]].drop(columns=["含む"]).reset_index(drop=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="最新版リスト")

        from openpyxl.utils import get_column_letter

        worksheet = writer.sheets["最新版リスト"]
        for idx, col in enumerate(df_export.columns):
            max_length = max(
                df_export[col].astype(str).str.len().max() if len(df_export) > 0 else 0,
                len(col),
            ) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)

    return buffer.getvalue()


def _export_compiled_csv(df: pd.DataFrame) -> bytes:
    """コンパイル済みリストを CSV にエクスポート

    Args:
        df: 「含む」列が True の行のみを出力対象とする DataFrame

    Returns:
        CSV ファイルのバイト列（UTF-8 BOM付き）
    """
    df_export = df[df["含む"]].drop(columns=["含む"]).reset_index(drop=True)
    return df_export.to_csv(index=False).encode("utf-8-sig")


def _render_compile_subtab() -> None:
    """サブタブ 3: 最新版リスト作成（LLM呼び出しなし、純粋なUI集約）"""

    st.markdown(
        "変更点調査と新規参入検出の結果を統合し、**最新版のプレイヤーリスト**を作成します。\n\n"
        "LLM呼び出しは行いません（コスト: 0円）。"
    )

    players = st.session_state.trend_players
    val_results = st.session_state.trend_val_results
    newcomer_candidates = st.session_state.trend_newcomer_candidates

    # ステータス表示
    st.markdown("### 📋 サブタブ完了状況")

    col1, col2, col3 = st.columns(3)
    with col1:
        if players:
            st.success(f"📂 Excel: {len(players)}件読み込み済み")
        else:
            st.warning("📂 Excel: 未アップロード")

    with col2:
        if val_results:
            st.success(f"📋 変更点調査: {len(val_results)}件完了")
        else:
            st.info("📋 変更点調査: 未実行")

    with col3:
        if newcomer_candidates:
            # 選択済み候補数をカウント
            selected_newcomers = [
                c for i, c in enumerate(newcomer_candidates)
                if st.session_state.get(f"trend_newcomer_select_{i}", False)
            ]
            st.success(
                f"🆕 新規参入検出: {len(newcomer_candidates)}件検出"
                f"（{len(selected_newcomers)}件選択中）"
            )
        else:
            st.info("🆕 新規参入検出: 未実行")

    st.divider()

    # Excel未アップロード時
    if not players:
        st.info("Excelファイルをアップロードしてから、このタブをご利用ください。")
        return

    # どちらも未実行の場合
    if not val_results and not newcomer_candidates:
        st.info(
            "変更点調査または新規参入検出を実行してから、このタブで統合リストを作成してください。\n\n"
            "- いずれか一方のみでも作成可能です\n"
            "- 両方実行するとより正確なリストになります"
        )
        return

    # 選択済み新規参入候補を抽出
    selected_newcomers = []
    for i, candidate in enumerate(newcomer_candidates):
        if st.session_state.get(f"trend_newcomer_select_{i}", False):
            selected_newcomers.append(candidate)

    # コンパイル実行
    df_compiled = _compile_final_list(players, val_results, selected_newcomers)

    if df_compiled.empty:
        st.warning("コンパイル結果が空です。全プレイヤーが撤退・統合となった可能性があります。")
        return

    # 統計サマリー
    st.markdown("### 📊 コンパイル結果サマリー")

    source_counts = df_compiled["ソース"].value_counts()
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("📋 合計", f"{len(df_compiled)}件")
    with col2:
        st.metric("📄 元リスト", f"{source_counts.get('元リスト', 0)}件")
    with col3:
        st.metric("✏️ 名称変更", f"{source_counts.get('名称変更', 0)}件")
    with col4:
        st.metric("🆕 新規参入", f"{source_counts.get('新規参入', 0)}件")

    # 除外されたプレイヤーの表示
    if val_results:
        excluded = [
            r for r in val_results
            if r.change_type in (ChangeType.WITHDRAWAL, ChangeType.MERGER)
        ]
        if excluded:
            with st.expander(f"🚫 除外されたプレイヤー（{len(excluded)}件）"):
                for r in excluded:
                    details = " / ".join(r.change_details) if r.change_details else ""
                    st.markdown(
                        f"- **{r.player_name_original}** — {r.change_type.value}"
                        + (f"（{details}）" if details else "")
                    )

    st.divider()

    # 編集可能なテーブル
    st.markdown("### ✏️ 最新版リスト（編集可能）")
    st.caption("「含む」チェックを外すとエクスポートから除外されます。")

    edited_df = st.data_editor(
        df_compiled,
        use_container_width=True,
        height=500,
        num_rows="fixed",
        column_config={
            "含む": st.column_config.CheckboxColumn("含む", default=True, width="small"),
            "プレイヤー名": st.column_config.TextColumn("プレイヤー名", width="medium"),
            "公式URL": st.column_config.TextColumn("公式URL", width="large"),
            "運営会社": st.column_config.TextColumn("運営会社", width="medium"),
            "ソース": st.column_config.TextColumn("ソース", width="small"),
            "備考": st.column_config.TextColumn("備考", width="large"),
        },
        key="trend_compile_editor",
    )

    # 選択件数
    included_count = edited_df["含む"].sum()
    st.info(f"エクスポート対象: {included_count}件 / {len(edited_df)}件")

    st.divider()

    # エクスポート
    st.subheader("📥 エクスポート")

    if included_count == 0:
        st.warning("エクスポート対象が0件です。「含む」チェックを確認してください。")
        return

    col1, col2 = st.columns(2)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    with col1:
        excel_data = _export_compiled_excel(edited_df)
        st.download_button(
            "📥 Excel ダウンロード（最新版リスト）",
            excel_data,
            f"player_list_latest_{timestamp}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="trend_compile_export_excel",
        )

    with col2:
        csv_data = _export_compiled_csv(edited_df)
        st.download_button(
            "📥 CSV ダウンロード",
            csv_data,
            f"player_list_latest_{timestamp}.csv",
            "text/csv",
            use_container_width=True,
            key="trend_compile_export_csv",
        )


# ====================================
# メインレンダリング
# ====================================
def render_player_trend_tab(industry: str, definition: str = "") -> None:
    """プレイヤーの最新動向タブのUIをレンダリング

    3つのサブタブを統合:
    1. 変更点調査 — 既存プレイヤーの撤退・統合・名称変更を検出
    2. 新規参入検出 — 既存リストにない新規参入候補を提案
    3. 最新版リスト作成 — 上記結果を統合した最新版リストを生成

    Args:
        industry: 業界名（例: "クレジットカード", "動画配信サービス"）
        definition: 業界定義テキスト（プレイヤーの判定基準）
    """
    _init_session_state()

    # 共通: Excel アップロード（サブタブの前に配置）
    _render_shared_upload()

    st.divider()

    # 3つのサブタブ
    tab_validation, tab_newcomer, tab_compile = st.tabs([
        "📋 変更点調査",
        "🆕 新規参入検出",
        "📊 最新版リスト作成",
    ])

    with tab_validation:
        _render_validation_subtab(industry, definition)

    with tab_newcomer:
        _render_newcomer_subtab(industry, definition)

    with tab_compile:
        _render_compile_subtab()
